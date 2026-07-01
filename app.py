"""
Employee Management System - Flask Backend (MongoDB Atlas)
===========================================================
Run with: python app.py
"""

# Load .env FIRST so all os.environ.get() calls throughout the app read correct values
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # dotenv optional; env vars may be set by the OS/host

from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, Blueprint, make_response, flash
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_talisman import Talisman
from flask_cors import CORS
from flask_wtf.csrf import CSRFProtect, CSRFError
from pymongo import MongoClient
import pymongo.errors
from bson.objectid import ObjectId
import certifi


import hashlib
import secrets
import os
import csv
import io
try:
    import dns.resolver
    dns.resolver.default_resolver = dns.resolver.Resolver(configure=False)
    dns.resolver.default_resolver.nameservers = ['8.8.8.8', '8.8.4.4', '1.1.1.1']
except Exception:
    pass
import calendar
import bcrypt
import logging
import json
import urllib.request
import urllib.parse
import sentry_sdk
from sentry_sdk.integrations.flask import FlaskIntegration
from datetime import datetime, date, timedelta
from functools import wraps
import pyotp
import qrcode
import base64
from io import BytesIO
from utils.mailer import send_verification_email, send_reset_email

import cloudinary
import cloudinary.uploader
import cloudinary.api

# Cloudinary config (will automatically use CLOUDINARY_URL env var if set)
cloudinary.config(secure=True)

# ─────────────────────────────────────────
# SECURITY & MONITORING INIT
# ─────────────────────────────────────────

# Initialize Sentry Error Monitoring (if DSN provided)
sentry_dsn = os.environ.get('SENTRY_DSN')
if sentry_dsn:
    sentry_sdk.init(
        dsn=sentry_dsn,
        integrations=[FlaskIntegration()],
        traces_sample_rate=0.1,
        # Scrub header keys and potentially sensitive info
        before_send=lambda event, hint: {
            **event,
            'request': {**event.get('request', {}), 'headers': '[REDACTED]'}
        } if event else None,
    )

app = Flask(__name__)

# Enforce SECRET_KEY crash on startup if missing
secret = os.environ.get('SECRET_KEY')
if not secret:
    raise RuntimeError("SECRET_KEY environment variable is not set. Refusing to start.")
app.secret_key = secret

# Structured JSON Logging
class JSONFormatter(logging.Formatter):
    def format(self, record):
        log_entry = {
            "level": record.levelname,
            "msg": record.getMessage(),
            "time": self.formatTime(record),
            "module": record.module,
        }
        if record.exc_info:
            log_entry["exc"] = self.formatException(record.exc_info)
        return json.dumps(log_entry)

app.logger.handlers.clear()
handler = logging.StreamHandler()
handler.setFormatter(JSONFormatter())
app.logger.addHandler(handler)
app.logger.setLevel(logging.INFO)

# Flask-Cors configuration
cors_origins_env = os.environ.get('CORS_ORIGINS', 'http://127.0.0.1:5000')
CORS(app,
     origins=cors_origins_env.split(','),
     supports_credentials=True,
     methods=['GET', 'POST', 'PATCH', 'DELETE', 'OPTIONS'])

# CSRF Protection Setup
csrf = CSRFProtect(app)

# Flask-Talisman Security Headers
is_prod = os.environ.get('FLASK_ENV', 'production') == 'production'
Talisman(app,
    force_https=is_prod,
    content_security_policy={
        'default-src': "'self'",
        'script-src': ["'self'", "'unsafe-inline'", "cdn.jsdelivr.net", "cdn.tailwindcss.com"],
        'style-src':  ["'self'", "'unsafe-inline'", "fonts.googleapis.com", "cdn.jsdelivr.net"],
        'font-src':   ["'self'", "fonts.gstatic.com", "cdn.jsdelivr.net"],
        'img-src':    ["'self'", "data:", "*"],
    },
    content_security_policy_nonce_in=[],
    frame_options='DENY',
    referrer_policy='strict-origin-when-cross-origin',
)

# Session Cookie Security
app.config.update(
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=is_prod,
)

# Request size limits (1MB max payload)
app.config['MAX_CONTENT_LENGTH'] = 1 * 1024 * 1024

@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    app.logger.warning(f"CSRF validation failed: {e.description}")
    if request.path.startswith('/api/') or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
        return jsonify({"success": False, "error": "CSRF token missing or invalid."}), 403
    return render_template("error.html", code=403, title="Forbidden", message="CSRF token missing or invalid. Please reload the page and try again."), 403

@app.errorhandler(403)
def forbidden_error(e):
    if request.path.startswith('/api/') or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
        return jsonify({"success": False, "error": "Forbidden"}), 403
    return render_template("error.html", code=403, title="Forbidden", message=str(e.description or "You do not have permission to access this resource.")), 403

@app.errorhandler(404)
def not_found_error(e):
    if request.path.startswith('/api/') or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
        return jsonify({"success": False, "error": "Not Found"}), 404
    return render_template("error.html", code=404, title="Not Found", message="The requested page could not be found."), 404

@app.errorhandler(500)
def internal_error(e):
    app.logger.error(f"Internal Server Error: {e}", exc_info=True)
    if request.path.startswith('/api/') or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
        return jsonify({"success": False, "error": "An internal server error occurred."}), 500
    return render_template("error.html", code=500, title="Internal Server Error", message="An unexpected error occurred on the server. Please try again later."), 500

@app.errorhandler(413)
def request_too_large(e):
    app.logger.warning("Payload size limit exceeded (413 error)")
    if request.path.startswith('/api/') or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
        return jsonify({
            "success": False,
            "error": {
                "code": "PAYLOAD_TOO_LARGE",
                "message": "Request payload exceeds 1MB limit."
            }
        }), 413
    return render_template("error.html", code=413, title="Payload Too Large", message="Request payload exceeds 1MB limit."), 413


# ─────────────────────────────────────────
# API V1 BLUEPRINT DECLARATION
# ─────────────────────────────────────────
api_v1 = Blueprint('api_v1', __name__, url_prefix='/api/v1')



DB_PATH = os.environ.get('DATABASE_PATH', os.path.join(os.path.dirname(__file__), 'database.db'))

# ─────────────────────────────────────────
# RATE LIMITER SETUP
# ─────────────────────────────────────────

limiter_storage = os.environ.get('REDIS_URL') or os.environ.get('REDIS_URI') or "memory://"
limiter = Limiter(
    key_func=get_remote_address,      # limit by IP address
    app=app,
    default_limits=["300 per day", "60 per hour"],  # global fallback
    storage_uri=limiter_storage,
)

@app.errorhandler(429)
def rate_limit_exceeded(e):
    """Return JSON for API calls, friendly HTML for browser requests."""
    if request.path.startswith('/api/') or request.is_json:
        return jsonify({
            'error': 'Too many requests',
            'message': str(e.description),
            'retry_after': e.retry_after if hasattr(e, 'retry_after') else 60
        }), 429
    # Browser-friendly 429 page
    retry = getattr(e, 'retry_after', 60)
    html = f"""<!DOCTYPE html>
<html><head><title>Too Many Requests — EMS</title>
<style>
  body{{font-family:'Plus Jakarta Sans',sans-serif;background:#0f172a;color:#f1f5f9;
       display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0;}}
  .box{{text-align:center;padding:48px;background:#1e293b;border-radius:20px;
        border:1px solid rgba(255,255,255,0.08);max-width:420px;}}
  .icon{{font-size:52px;margin-bottom:16px;}}
  h1{{font-size:22px;font-weight:800;margin:0 0 10px;}}
  p{{color:#94a3b8;font-size:14px;line-height:1.7;}}
  .badge{{display:inline-block;margin-top:20px;padding:8px 20px;background:#3b82f6;
           border-radius:10px;font-size:13px;font-weight:700;color:#fff;}}
  a{{color:#3b82f6;text-decoration:none;}}
</style></head>
<body><div class="box">
  <div class="icon">🛑</div>
  <h1>Too Many Requests</h1>
  <p>{e.description}</p>
  <p>Please wait <strong>{retry} seconds</strong> before trying again.</p>
  <a href="/" class="badge">← Go Back</a>
</div></body></html>"""
    return html, 429


# ─────────────────────────────────────────
# DATABASE SETUP (MONGODB ATLAS)
# ─────────────────────────────────────────

mongo_uri = os.environ.get('MONGO_URI')
if not mongo_uri:
    raise RuntimeError("MONGO_URI environment variable is not set. Refusing to start.")

import time
from pymongo import monitoring
from flask import g, template_rendered, before_render_template, has_app_context, has_request_context

ENABLE_PROFILING = os.environ.get('ENABLE_PROFILING') == 'true' or os.environ.get('FLASK_ENV') == 'development'

class MongoProfiler(monitoring.CommandListener):
    def started(self, event):
        if ENABLE_PROFILING and has_app_context() and has_request_context():
            if not hasattr(g, 'mongo_queries'):
                g.mongo_queries = []
            
            coll_name = "unknown"
            if event.command_name in event.command:
                coll_name = event.command.get(event.command_name, "unknown")
            elif event.command_name == 'insert':
                coll_name = event.command.get('insert', "unknown")
            elif event.command_name == 'update':
                coll_name = event.command.get('update', "unknown")
            elif event.command_name == 'delete':
                coll_name = event.command.get('delete', "unknown")
            elif event.command_name == 'aggregate':
                coll_name = event.command.get('aggregate', "unknown")

            doc_count = 0
            if 'documents' in event.command:
                doc_count = len(event.command['documents'])
            elif 'updates' in event.command:
                doc_count = len(event.command['updates'])

            g.mongo_queries.append({
                'request_id': event.request_id,
                'name': event.command_name,
                'collection': str(coll_name),
                'doc_count': doc_count,
                'start_time': time.time(),
                'duration': 0
            })

    def succeeded(self, event):
        if ENABLE_PROFILING and has_app_context() and has_request_context():
            duration_ms = event.duration_micros / 1000.0
            if hasattr(g, 'mongo_time'):
                g.mongo_time += duration_ms
            if hasattr(g, 'mongo_queries'):
                for q in reversed(g.mongo_queries):
                    if q.get('request_id') == event.request_id:
                        q['duration'] = duration_ms
                        if event.reply and 'cursor' in event.reply and 'firstBatch' in event.reply['cursor']:
                            q['doc_count'] = len(event.reply['cursor']['firstBatch'])
                        elif event.reply and 'n' in event.reply:
                            q['doc_count'] = event.reply['n']
                        break

    def failed(self, event):
        if ENABLE_PROFILING and has_app_context() and has_request_context():
            duration_ms = event.duration_micros / 1000.0
            if hasattr(g, 'mongo_time'):
                g.mongo_time += duration_ms

if ENABLE_PROFILING:
    monitoring.register(MongoProfiler())

def before_render(sender, template, context, **extra):
    if ENABLE_PROFILING:
        g.render_start_time = time.time()

def after_render(sender, template, context, **extra):
    if ENABLE_PROFILING and hasattr(g, 'render_start_time'):
        duration = (time.time() - g.render_start_time) * 1000.0
        g.render_time = getattr(g, 'render_time', 0.0) + duration

if ENABLE_PROFILING:
    before_render_template.connect(before_render, app)
    template_rendered.connect(after_render, app)

@app.before_request
def start_timer():
    if ENABLE_PROFILING:
        g.start_time = time.time()
        g.mongo_time = 0.0
        g.render_time = 0.0
        g.mongo_queries = []

@app.after_request
def log_request_timing(response):
    if ENABLE_PROFILING and hasattr(g, 'start_time'):
        total_time = (time.time() - g.start_time) * 1000.0
        route_name = request.endpoint or "Unknown"
        queries = getattr(g, 'mongo_queries', [])
        
        log_record = {
            "route": route_name,
            "path": request.path,
            "method": request.method,
            "total_time_ms": round(total_time, 2),
            "mongo_time_ms": round(getattr(g, 'mongo_time', 0.0), 2),
            "render_time_ms": round(getattr(g, 'render_time', 0.0), 2),
            "query_count": len(queries),
            "queries": [
                {
                    "name": q.get('name'),
                    "collection": q.get('collection'),
                    "duration_ms": round(q.get('duration', 0.0), 2),
                    "doc_count": q.get('doc_count', 0)
                } for q in queries
            ]
        }
        
        app.logger.info(f"PROFILING: {route_name} | Total: {log_record['total_time_ms']}ms | Mongo: {log_record['mongo_time_ms']}ms | Render: {log_record['render_time_ms']}ms | Queries: {log_record['query_count']}")
        
        os.makedirs("scratch", exist_ok=True)
        try:
            with open("scratch/profiling.jsonl", "a") as f:
                f.write(json.dumps(log_record) + "\n")
        except Exception:
            pass
            
    return response

try:
    mongo_client = MongoClient(mongo_uri, tlsCAFile=certifi.where())
    db = mongo_client.get_default_database()
except Exception as e:
    try:
        db = mongo_client['ems_db']
    except Exception:
        raise RuntimeError(f"Failed to initialize MongoDB connection: {e}")


def init_db():
    """Initialize MongoDB collections and create unique indexes."""
    try:
        db.users.create_index("email", unique=True)
        db.users.create_index("username", unique=True)
        db.employees.create_index([("user_id", 1), ("name", 1)])
        
        # Performance: Compound Indexes for high-volume queries
        db.part_time_workers.create_index([("user_id", 1), ("name", 1)])
        db.part_time_work_logs.create_index([("user_id", 1), ("worker_id", 1), ("working_date", -1)])
        db.part_time_work_logs.create_index([("worker_id", 1), ("working_date", -1)])
        db.advance_payments.create_index([("work_log_id", 1)])
        db.advance_payments.create_index([("user_id", 1), ("payment_date", -1)])
        db.advance_payments.create_index([("worker_id", 1)])
        db.attendance.create_index([("user_id", 1), ("date", -1)])
        db.attendance.create_index([("emp_id", 1), ("date", 1)])
        db.salary_records.create_index([("user_id", 1), ("month", -1)])
        db.salary_records.create_index([("emp_id", 1), ("month", 1)])
        db.salary_advance_payments.create_index([("salary_record_id", 1)])
        
        # Ensure counters exist, but do not override if they already exist
        db.counters.update_one(
            {"_id": "employee_id_seq"},
            {"$setOnInsert": {"seq": 0}},
            upsert=True
        )
        db.counters.update_one(
            {"_id": "worker_id_seq"},
            {"$setOnInsert": {"seq": 0}},
            upsert=True
        )
        app.logger.info("MongoDB indexes and counters verified successfully.")
    except Exception as e:
        app.logger.error(f"Error initializing MongoDB: {e}", exc_info=True)

def get_next_sequence(sequence_name):
    """Atomically increment and return the next sequence number."""
    result = db.counters.find_one_and_update(
        {"_id": sequence_name},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    return result["seq"]

def generate_employee_id():
    seq = get_next_sequence("employee_id_seq")
    return f"EMP{seq:03d}"

def generate_worker_id():
    seq = get_next_sequence("worker_id_seq")
    return f"WRK{seq:03d}"


try:
    init_db()
except Exception as e:
    app.logger.error(f"Import time init_db error: {e}")
    pass


# ─────────────────────────────────────────
# HEALTH & READY ENDPOINTS
# ─────────────────────────────────────────

@app.route('/health')
def health():
    """Liveness probe returning simple OK status."""
    return jsonify({"status": "ok"}), 200


@app.route('/ready')
def ready():
    """Readiness probe checking MongoDB connection status."""
    try:
        db.client.admin.command('ping')
        return jsonify({"status": "ready", "database": "connected"}), 200
    except Exception as e:
        app.logger.error(f"Readiness probe failed: {e}")
        return jsonify({"status": "not_ready", "error": str(e)}), 503


# ─────────────────────────────────────────

# DOCUMENT SERIALIZATION HELPERS
# ─────────────────────────────────────────

def serialize_doc(doc):
    """Adds a string 'id' field and safely casts ObjectIds to strings."""
    if not doc:
        return None
    doc = dict(doc)
    if '_id' in doc:
        doc['id'] = str(doc['_id'])
        del doc['_id']
    for k, v in doc.items():
        if isinstance(v, ObjectId):
            doc[k] = str(v)
    return doc


def serialize_docs(docs):
    """Serialize a list of MongoDB documents."""
    return [serialize_doc(d) for d in docs]


def safe_object_id(val):
    """Safely cast value to ObjectId, returning None if invalid."""
    if not val:
        return None
    try:
        return ObjectId(str(val))
    except Exception:
        return None




# ─────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────

def hash_password(password):
    """Hash password using bcrypt (rounds=12)."""
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt(12)).decode('utf-8')


def verify_and_migrate_password(user_id, password, stored_hash):
    """
    Verify password using bcrypt. If stored_hash is a legacy SHA-256 hash
    (64 chars, not starting with '$'), verify using SHA-256, and if successful,
    automatically upgrade/re-hash using bcrypt and update the database.
    """
    is_legacy = len(stored_hash) == 64 and not stored_hash.startswith('$')
    
    if is_legacy:
        sha256_hash = hashlib.sha256(password.encode('utf-8')).hexdigest()
        if sha256_hash == stored_hash:
            new_bcrypt_hash = hash_password(password)
            db.users.update_one({"_id": ObjectId(user_id)}, {"$set": {"password": new_bcrypt_hash}})
            app.logger.info("Automatically migrated user password from SHA-256 to bcrypt", extra={"user_id": str(user_id)})
            return True
        return False
    else:
        try:
            return bcrypt.checkpw(password.encode('utf-8'), stored_hash.encode('utf-8'))
        except Exception as e:
            app.logger.error("Bcrypt check failed", exc_info=True)
            return False



def login_required(f):
    """Decorator to protect routes that need login."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not get_current_user_id():
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def get_current_user_id():
    uid = session.get('user_id')
    if not uid:
        return None
    try:
        # Validate that uid is a valid hex string of 24 characters (ObjectId)
        ObjectId(uid)
        return uid
    except Exception:
        # Force logout/clear of legacy SQLite integer user_id or corrupt session
        session.clear()
        return None



# ─────────────────────────────────────────
# AUTH ROUTES
# ─────────────────────────────────────────

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/signup', methods=['GET', 'POST'])
@limiter.limit("5 per hour")            # 5 signups/hour per IP
def signup():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email    = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        confirm  = request.form.get('confirm_password', '')

        # Validation
        if not email or not password:
            return render_template('signup.html', error='All fields are required.')
        if '@' not in email or '.' not in email.split('@')[-1]:
            return render_template('signup.html', error='Please enter a valid email address.')
        if len(password) < 8:
            return render_template('signup.html', error='Password must be at least 8 characters.')
        if password != confirm:
            return render_template('signup.html', error='Passwords do not match.')

        username = email.split('@')[0]
        if db.users.find_one({"email": email}):
            return render_template('signup.html', error='An account with this email already exists.')
        if db.users.find_one({"username": username}):
            return render_template('signup.html', error='An account with this username already exists.')

        verify_token = secrets.token_urlsafe(32)
        try:
            db.users.insert_one({
                "username": username,
                "email": email,
                "password": hash_password(password),
                "otp_secret": None,
                "otp_enabled": False,
                "is_verified": False,
                "verify_token": verify_token,
                "reset_token": None,
                "reset_token_expiry": None,
                "created_at": datetime.utcnow().isoformat()
            })
        except pymongo.errors.DuplicateKeyError as e:
            # Fallback catch in case of race conditions
            app.logger.warning(f"Race condition DuplicateKeyError during signup: {e}")
            return render_template('signup.html', error='An account with this email or username already exists.')


        send_verification_email(email, verify_token)
        return render_template('signup.html', sent=True, email=email)

    return render_template('signup.html')


@app.route('/verify/<token>')
def verify_email(token):
    user = db.users.find_one({"verify_token": token, "is_verified": False})
    if not user:
        return render_template('verify_email.html', status='invalid')
    
    db.users.update_one(
        {"_id": user['_id']},
        {"$set": {"is_verified": True, "verify_token": None}}
    )
    return render_template('verify_email.html', status='success')


@app.route('/resend-verification', methods=['POST'])
@limiter.limit("3 per hour")
def resend_verification():
    email = request.form.get('email', '').strip().lower()
    if not email:
        return redirect(url_for('login'))
    
    user = db.users.find_one({"email": email, "is_verified": False})
    if user:
        new_token = secrets.token_urlsafe(32)
        db.users.update_one(
            {"_id": user['_id']},
            {"$set": {"verify_token": new_token}}
        )
        send_verification_email(email, new_token)
    return render_template('login.html', success='Verification email resent! Check your inbox.')


# ── GOOGLE OAUTH ROUTES ───────────────────────────────────

@app.route('/auth/google')
@limiter.limit("20 per minute")
def login_google():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
        
    client_id = os.environ.get('GOOGLE_CLIENT_ID')
    if not client_id:
        return render_template('login.html', error="Google Client ID is not configured in the environment.")

    # Generate a secure state token for CSRF protection
    state = secrets.token_urlsafe(16)
    session['oauth_state'] = state

    # Construct the redirect URI: prioritize static env override, otherwise build dynamically (works on Vercel and localhost)
    redirect_uri = os.environ.get('GOOGLE_CALLBACK_URL')
    if not redirect_uri:
        scheme = 'https' if 'vercel' in request.host or 'ems' in request.host else 'http'
        host = request.host.replace('127.0.0.1', 'localhost')
        redirect_uri = f"{scheme}://{host}/auth/google/callback"

    params = {
        'client_id': client_id,
        'redirect_uri': redirect_uri,
        'response_type': 'code',
        'scope': 'openid email profile',
        'state': state
    }
    auth_url = "https://accounts.google.com/o/oauth2/v2/auth?" + urllib.parse.urlencode(params)
    return redirect(auth_url)


@app.route('/auth/google/callback')
@limiter.limit("20 per minute")
def auth_google_callback():
    # 1. State verification to prevent CSRF attacks
    state = request.args.get('state')
    saved_state = session.pop('oauth_state', None)
    if not state or state != saved_state:
        return render_template('login.html', error="Invalid CSRF state. Please try logging in again.")

    code = request.args.get('code')
    if not code:
        return render_template('login.html', error="Authorization failed. Google did not return a code.")

    client_id = os.environ.get('GOOGLE_CLIENT_ID')
    client_secret = os.environ.get('GOOGLE_CLIENT_SECRET')
    
    redirect_uri = os.environ.get('GOOGLE_CALLBACK_URL')
    if not redirect_uri:
        scheme = 'https' if 'vercel' in request.host or 'ems' in request.host else 'http'
        host = request.host.replace('127.0.0.1', 'localhost')
        redirect_uri = f"{scheme}://{host}/auth/google/callback"

    # 2. Exchange authorization code for Access & ID tokens
    token_url = "https://oauth2.googleapis.com/token"
    payload = urllib.parse.urlencode({
        'code': code,
        'client_id': client_id,
        'client_secret': client_secret,
        'redirect_uri': redirect_uri,
        'grant_type': 'authorization_code'
    }).encode('utf-8')

    token_req = urllib.request.Request(token_url, data=payload, method='POST')
    token_req.add_header('Content-Type', 'application/x-www-form-urlencoded')

    try:
        with urllib.request.urlopen(token_req) as response:
            token_data = json.loads(response.read().decode('utf-8'))
            access_token = token_data.get('access_token')
    except Exception as e:
        app.logger.error(f"Google Token Exchange Error: {e}")
        return render_template('login.html', error="Failed to authenticate with Google. Token exchange failed.")

    # 3. Retrieve user profile details from Google UserInfo API
    user_info_url = "https://www.googleapis.com/oauth2/v3/userinfo"
    user_req = urllib.request.Request(user_info_url)
    user_req.add_header('Authorization', f'Bearer {access_token}')

    try:
        with urllib.request.urlopen(user_req) as response:
            user_info = json.loads(response.read().decode('utf-8'))
    except Exception as e:
        app.logger.error(f"Google User Info Fetch Error: {e}")
        return render_template('login.html', error="Failed to retrieve Google profile information.")

    google_id = user_info.get('sub')
    email = user_info.get('email', '').strip().lower()
    name = user_info.get('name')
    avatar = user_info.get('picture')

    if not email:
        return render_template('login.html', error="Google account did not provide a primary email address.")

    username = email.split('@')[0]

    # 4. Synchronize user in MongoDB Atlas
    # Check if a user with this google_id already exists
    user = db.users.find_one({"google_id": google_id})

    if not user:
        # Fallback: check by email in case of manual pre-existing signups
        user = db.users.find_one({"email": email})
        
        if user:
            # Upgrade local account to support Google OAuth
            db.users.update_one(
                {"_id": user['_id']},
                {"$set": {
                    "google_id": google_id,
                    "avatar": avatar,
                    "is_verified": True,   # Google email is already verified
                    "last_login": datetime.utcnow().isoformat()
                }}
            )
            user = db.users.find_one({"_id": user['_id']})
        else:
            # Register a brand new Google user
            new_user = {
                "username": username,
                "email": email,
                "password": hash_password(secrets.token_urlsafe(16)),  # secure placeholder password
                "google_id": google_id,
                "avatar": avatar,
                "is_verified": True,
                "verify_token": None,
                "otp_secret": None,
                "otp_enabled": False,
                "reset_token": None,
                "reset_token_expiry": None,
                "created_at": datetime.utcnow().isoformat(),
                "last_login": datetime.utcnow().isoformat()
            }
            res = db.users.insert_one(new_user)
            user = db.users.find_one({"_id": res.inserted_id})
    else:
        # User exists, update avatar and login time
        db.users.update_one(
            {"_id": user['_id']},
            {"$set": {
                "avatar": avatar,
                "last_login": datetime.utcnow().isoformat()
            }}
        )

    # 5. Establish secure authenticated session cookie
    session.clear()
    session['user_id'] = str(user['_id'])
    session['username'] = user['username']
    session['email'] = user['email']
    if user.get('avatar'):
        session['avatar'] = user.get('avatar')

    app.logger.info("Google OAuth login successful", extra={"user_id": str(user['_id'])})
    return redirect(url_for('dashboard'))



@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute")         # 10 login attempts/min — brute-force protection
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    success = request.args.get('success')
    if request.method == 'POST':
        email    = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        remember = request.form.get('remember_me') == 'on'

        if not email or not password:
            return render_template('login.html', error='Email and password are required.')

        user = db.users.find_one({"email": email})
        
        if user and verify_and_migrate_password(user['_id'], password, user['password']):
            # Password verification succeeded
            app.logger.info("User login successful", extra={"user_id": str(user['_id'])})
        else:
            user = None

        if not user:
            return render_template('login.html', error='Invalid email or password.')

        # Block unverified users
        if not user.get('is_verified', True):
            return render_template('login.html',
                error='Please verify your email before logging in.',
                unverified_email=email)

        # Prevent Session Fixation
        session.clear()
        session['user_id']  = str(user['_id'])
        session['username'] = user['username']
        session['email'] = user['email']
        if user.get('avatar'):
            session['avatar'] = user.get('avatar')
        session['last_login'] = datetime.now().strftime('%b %d, %Y at %I:%M %p')
        if remember:
            from flask import current_app
            session.permanent = True
            current_app.permanent_session_lifetime = timedelta(days=30)

        return redirect(url_for('dashboard'))

    return render_template('login.html', success=success)


@app.route('/forgot-password', methods=['GET', 'POST'])
@limiter.limit("5 per hour")
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        if email:
            user = db.users.find_one({"email": email})
            if user:
                token  = secrets.token_urlsafe(32)
                expiry = (datetime.utcnow() + timedelta(hours=1)).isoformat()
                db.users.update_one(
                    {"_id": user['_id']},
                    {"$set": {"reset_token": token, "reset_token_expiry": expiry}}
                )
                send_reset_email(email, token)
        # Always show success (prevents user enumeration)
        return render_template('forgot_password.html', sent=True, email=email)
    return render_template('forgot_password.html')


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    user = db.users.find_one({"reset_token": token})

    # Token not found
    if not user:
        return render_template('reset_password.html', status='invalid')

    # Token expired
    if user.get('reset_token_expiry'):
        try:
            expiry = datetime.fromisoformat(user['reset_token_expiry'])
            if datetime.utcnow() > expiry:
                return render_template('reset_password.html', status='expired')
        except ValueError:
            return render_template('reset_password.html', status='invalid')

    if request.method == 'POST':
        password = request.form.get('password', '')
        confirm  = request.form.get('confirm_password', '')
        if len(password) < 8:
            return render_template('reset_password.html', status='form', token=token,
                                   error='Password must be at least 8 characters.')
        if password != confirm:
            return render_template('reset_password.html', status='form', token=token,
                                   error='Passwords do not match.')
        
        db.users.update_one(
            {"_id": user['_id']},
            {"$set": {
                "password": hash_password(password),
                "reset_token": None,
                "reset_token_expiry": None
            }}
        )
        return render_template('reset_password.html', status='success')

    # GET — show the form
    return render_template('reset_password.html', status='form', token=token)




@api_v1.route('/profile-info')
@login_required
def api_profile_info():
    """Return current user info as JSON for the profile popup."""
    user = db.users.find_one({"_id": ObjectId(get_current_user_id())})
    user = serialize_doc(user)
    return jsonify({
        'username':   user['username'] if user else session.get('username', ''),
        'email':      user['email']    if user else '',
        'last_login': session.get('last_login', 'Unknown'),
        'avatar':     user.get('avatar') if user else session.get('avatar', ''),
        'profile_image_url': user.get('profile_image_url') if user else '',
        'is_google':  bool(user.get('google_id')) if user else False
    })



@api_v1.route('/change-password', methods=['POST'])
@login_required
@limiter.limit("5 per hour")
def api_change_password():
    """Change password from the profile popup (JSON API)."""
    data        = request.get_json(silent=True) or {}
    current_pw  = data.get('current_password', '')
    new_pw      = data.get('new_password', '')
    confirm_pw  = data.get('confirm_password', '')

    if not current_pw or not new_pw or not confirm_pw:
        return jsonify({'ok': False, 'error': 'All fields are required.'}), 400
    if len(new_pw) < 8:
        return jsonify({'ok': False, 'error': 'Password must be at least 8 characters.'}), 400
    if new_pw != confirm_pw:
        return jsonify({'ok': False, 'error': 'New passwords do not match.'}), 400

    user = db.users.find_one({"_id": ObjectId(get_current_user_id())})
    
    if not user or not verify_and_migrate_password(user['_id'], current_pw, user['password']):
        return jsonify({'ok': False, 'error': 'Current password is incorrect.'}), 403

    db.users.update_one(
        {"_id": user['_id']},
        {"$set": {"password": hash_password(new_pw)}}
    )
    return jsonify({'ok': True, 'message': 'Password changed successfully.'})


@app.route('/logout')
def logout():
    session.clear()
    response = make_response(redirect(url_for('login')))
    response.set_cookie(app.config.get('SESSION_COOKIE_NAME', 'session'), '', expires=0)
    return response





# ─────────────────────────────────────────
# DASHBOARD CACHE
# ─────────────────────────────────────────
from datetime import datetime, timedelta

DASHBOARD_CACHE = {}

def get_dashboard_cache(uid):
    cache = DASHBOARD_CACHE.get(uid)
    if cache and datetime.now() - cache['time'] < timedelta(minutes=5):
        return cache['data']
    return None

def set_dashboard_cache(uid, data):
    DASHBOARD_CACHE[uid] = {'time': datetime.now(), 'data': data}

def invalidate_dashboard_cache(uid):
    if uid in DASHBOARD_CACHE:
        del DASHBOARD_CACHE[uid]

# ─────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────

@app.route('/dashboard')
@limiter.limit("120 per minute")        # generous — it's a read-only page
@login_required
def dashboard():
    uid = get_current_user_id()

    cached_data = get_dashboard_cache(uid)
    if cached_data:
        return render_template('dashboard.html', **cached_data)

    emp_pipeline = [
        {"$match": {"user_id": ObjectId(uid)}},
        {"$group": {
            "_id": None,
            "total_emp": {"$sum": 1},
            "avg_salary": {"$avg": "$salary"},
            "avg_age": {"$avg": "$age"},
            "total_hrs": {"$sum": "$working_hours"},
            "emp_ids": {"$push": "$_id"}
        }}
    ]
    emp_agg = list(db.employees.aggregate(emp_pipeline))
    
    total_emp = 0
    avg_salary = 0.0
    avg_age = 0.0
    total_hrs = 0.0
    emp_ids = []
    
    if emp_agg:
        data = emp_agg[0]
        total_emp = data.get('total_emp', 0)
        avg_salary = round(data.get('avg_salary') or 0.0, 2)
        avg_age = round(data.get('avg_age') or 0.0, 1)
        total_hrs = data.get('total_hrs', 0.0)
        emp_ids = data.get('emp_ids', [])

    present_count = 0
    absent_count = 0
    salary_data = []
    recent_att = []
    
    if emp_ids:
        emps = list(db.employees.find({"_id": {"$in": emp_ids}}, {"name": 1, "salary": 1}))
        salary_data = [{'name': e['name'], 'salary': e.get('salary', 0)} for e in emps]
        
        absent_count = db.attendance.count_documents({"emp_id": {"$in": emp_ids}, "status": "Absent"})
        present_count = db.attendance.count_documents({"emp_id": {"$in": emp_ids}, "status": "Present"})

        recent_rows = list(db.attendance.find({"emp_id": {"$in": emp_ids}}).sort("date", -1).limit(10))
        emp_dict = {e['_id']: e['name'] for e in emps}
        for r in recent_rows:
            recent_att.append({
                'name': emp_dict.get(r['emp_id'], 'Unknown'),
                'date': r['date'],
                'status': r['status']
            })

    today = date.today().isoformat()
    
    template_data = {
        'total_emp': total_emp,
        'avg_salary': avg_salary,
        'avg_age': avg_age,
        'total_hrs': total_hrs,
        'present_count': present_count,
        'absent_count': absent_count,
        'salary_data': salary_data,
        'recent_att': recent_att,
        'today': today
    }
    
    set_dashboard_cache(uid, template_data)
    
    return render_template('dashboard.html', **template_data)


# ─────────────────────────────────────────
# EMPLOYEES
# ─────────────────────────────────────────

@app.route('/employees')
@limiter.limit("120 per minute")
@login_required
def employees():
    uid = get_current_user_id()
    search = request.args.get('q', '').strip()
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 25))
    except ValueError:
        page = 1
        limit = 25
        
    skip = (page - 1) * limit

    query = {"user_id": ObjectId(uid)}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"employee_id": {"$regex": search, "$options": "i"}}
        ]

    total_records = db.employees.count_documents(query)
    total_pages = (total_records + limit - 1) // limit if limit > 0 else 1
    
    emps = list(db.employees.find(query).sort("name", 1).skip(skip).limit(limit))
    emps = serialize_docs(emps)
    
    # Filter out employees who haven't joined yet by selected_date
    filtered_emps = []
    for e in emps:
        jd = e.get('joining_date')
        if jd and selected_date < jd:
            continue
        filtered_emps.append(e)
    emps = filtered_emps
    
    
    return render_template('employees.html', 
                           employees=emps, 
                           search=search,
                           page=page,
                           limit=limit,
                           total_pages=total_pages,
                           total_records=total_records)


@app.route('/employees/add', methods=['POST'])
@limiter.limit("30 per minute")         # prevent rapid bulk employee creation
@login_required
def add_employee():
    uid = get_current_user_id()
    invalidate_dashboard_cache(uid)
    name = request.form.get('name', '').strip()
    if not name:
        redirect_to = request.form.get('redirect_to')
        if redirect_to == 'part_time':
            return redirect(url_for('part_time'))
        return redirect(url_for('employees'))

    phone = request.form.get('phone', '').strip()
    joining_date = request.form.get('joining_date', '').strip()
    if not joining_date:
        flash("Joining date is required.", "error")
        return redirect(url_for('employees'))
    
    gender = request.form.get('gender', '')

    try:
        age = int(request.form.get('age') or 0)
    except (TypeError, ValueError):
        age = 0
    age = max(age, 0)

    try:
        salary = float(request.form.get('salary') or 0.0)
    except (TypeError, ValueError):
        salary = 0.0
    salary = max(salary, 0.0)

    try:
        leaves = int(request.form.get('leaves') or 0)
    except (TypeError, ValueError):
        leaves = 0
    leaves = max(leaves, 0)

    try:
        hours = float(request.form.get('working_hours') or 40.0)
    except (TypeError, ValueError):
        hours = 40.0
    hours = max(hours, 0.0)

    profile_image_url = ""
    if 'profile_image' in request.files and request.files['profile_image'].filename != '':
        try:
            p_url = upload_avatar_to_cloudinary(request.files['profile_image'])
            if p_url: profile_image_url = p_url
        except Exception as e:
            flash(str(e), "error")

    # Generate unique employee ID
    emp_id_str = generate_employee_id()

    db.employees.insert_one({
        "employee_id": emp_id_str,
        "name": name,
        "phone": phone,
        "joining_date": joining_date,
        "age": age,
        "gender": gender,
        "salary": salary,
        "leaves": leaves,
        "working_hours": hours,
        "profile_image_url": profile_image_url,
        "user_id": ObjectId(uid),
        "created_at": datetime.utcnow().isoformat()
    })


    redirect_to = request.form.get('redirect_to')
    if redirect_to == 'part_time':
        return redirect(url_for('part_time'))
    return redirect(url_for('employees'))


@app.route('/employees/edit/<emp_id>', methods=['GET', 'POST'])
@limiter.limit("30 per minute")
@login_required
def edit_employee(emp_id):
    uid = get_current_user_id()
    invalidate_dashboard_cache(uid)
    emp_id_obj = safe_object_id(emp_id)
    if not emp_id_obj:
        return redirect(url_for('employees'))

    if request.method == 'POST':
        try:
            age = int(request.form.get('age') or 0)
        except (TypeError, ValueError):
            age = 0
        age = max(age, 0)

        try:
            salary = float(request.form.get('salary') or 0.0)
        except (TypeError, ValueError):
            salary = 0.0
        salary = max(salary, 0.0)

        try:
            leaves = int(request.form.get('leaves') or 0)
        except (TypeError, ValueError):
            leaves = 0
        leaves = max(leaves, 0)

        try:
            hours = float(request.form.get('working_hours') or 40.0)
        except (TypeError, ValueError):
            hours = 40.0
        hours = max(hours, 0.0)

        joining_date = request.form.get('joining_date', '').strip()
        if not joining_date:
            flash("Joining date is required.", "error")
            return redirect(url_for('edit_employee', emp_id=emp_id))
            
        update_data = {
            "name": request.form.get('name'),
            "phone": request.form.get('phone'),
            "joining_date": joining_date,
            "age": age,
            "gender": request.form.get('gender'),
            "salary": salary,
            "leaves": leaves,
            "working_hours": hours
        }

        if 'profile_image' in request.files and request.files['profile_image'].filename != '':
            try:
                p_url = upload_avatar_to_cloudinary(request.files['profile_image'])
                if p_url: update_data['profile_image_url'] = p_url
            except Exception as e:
                flash(str(e), "error")

        db.employees.update_one(
            {"_id": emp_id_obj, "user_id": ObjectId(uid)},
            {"$set": update_data}
        )
        return redirect(url_for('employees'))

    emp = db.employees.find_one({"_id": emp_id_obj, "user_id": ObjectId(uid)})
    emp = serialize_doc(emp)
    if not emp:
        return redirect(url_for('employees'))
    return render_template('edit_employee.html', emp=emp)



@app.route('/employees/delete/<emp_id>', methods=['POST'])
@limiter.limit("20 per minute")         # prevent rapid deletion attacks
@login_required
def delete_employee(emp_id):
    uid = get_current_user_id()
    invalidate_dashboard_cache(uid)
    emp_id_obj = safe_object_id(emp_id)
    app.logger.info(f"Deleting employee: {emp_id}")
    print(f"Deleting employee: {emp_id}")
    
    if not emp_id_obj:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
            return jsonify({"success": False, "error": "Invalid employee ID"}), 400
        return redirect(url_for('employees'))

    # Verify employee belongs to current user before deleting anything
    emp = db.employees.find_one({"_id": emp_id_obj, "user_id": ObjectId(uid)})
    if not emp:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
            return jsonify({"success": False, "error": "Employee not found or access denied"}), 404
        return redirect(url_for('employees'))

    db.attendance.delete_many({"emp_id": emp_id_obj})
    db.salary_records.delete_many({"emp_id": emp_id_obj})
    db.salary_advance_payments.delete_many({"emp_id": emp_id_obj})
    result = db.employees.delete_one({"_id": emp_id_obj, "user_id": ObjectId(uid)})
    app.logger.info(f"Deleted count: {result.deleted_count}")
    print(f"Deleted count: {result.deleted_count}")
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
        return jsonify({"success": True, "message": "Employee deleted successfully", "deleted_count": result.deleted_count})
    return redirect(url_for('employees'))



# ─────────────────────────────────────────
# ATTENDANCE
# ─────────────────────────────────────────

@app.route('/attendance')
@limiter.limit("120 per minute")
@login_required
def attendance():
    uid = get_current_user_id()
    selected_date = request.args.get('date', date.today().isoformat())

    search = request.args.get('q', '').strip()
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 25))
    except ValueError:
        page = 1
        limit = 25
        
    skip = (page - 1) * limit
    
    query = {"user_id": ObjectId(uid)}
    if search:
        query["name"] = {"$regex": search, "$options": "i"}

    total_records = db.employees.count_documents(query)
    total_pages = (total_records + limit - 1) // limit if limit > 0 else 1

    emps = list(db.employees.find(query).sort("name", 1).skip(skip).limit(limit))
    emps = serialize_docs(emps)
    
    # Filter out employees who haven't joined yet by selected_date
    filtered_emps = []
    for e in emps:
        jd = e.get('joining_date')
        if jd and selected_date < jd:
            continue
        filtered_emps.append(e)
    emps = filtered_emps
    
    att_map = {e['id']: 'Present' for e in emps}

    if emps:
        emp_ids = [ObjectId(e['id']) for e in emps]
        records = list(db.attendance.find({
            "date": selected_date,
            "emp_id": {"$in": emp_ids}
        }))
        for r in records:
            att_map[str(r['emp_id'])] = r['status']

    return render_template(
        'attendance.html',
        employees=emps,
        att_map=att_map,
        selected_date=selected_date,
        today=date.today().isoformat(),
        search=search,
        page=page,
        limit=limit,
        total_pages=total_pages,
        total_records=total_records
    )


@app.route('/attendance/mark', methods=['POST'])
@limiter.limit("120 per minute")        # high limit — real-time marking via JS clicks
@login_required
def mark_attendance():
    invalidate_dashboard_cache(get_current_user_id())
    data = request.get_json() or {}
    emp_id = data.get('emp_id')
    att_date = data.get('date')
    status = data.get('status')
    leave_reason = data.get('leave_reason')
    leave_note = data.get('leave_note')

    if not all([emp_id, att_date, status]):
        return jsonify({'success': False, 'message': 'Missing parameters'}), 400

    emp_id_obj = safe_object_id(emp_id)
    if not emp_id_obj:
        return jsonify({'success': False, 'message': 'Invalid Employee ID'}), 400

    uid = get_current_user_id()
    # Verify employee belongs to current user
    emp = db.employees.find_one({"_id": emp_id_obj, "user_id": ObjectId(uid)})
    if not emp:
        return jsonify({'success': False, 'message': 'Unauthorized or Employee not found'}), 403

    try:
        if status == 'Present':
            db.attendance.delete_one({"emp_id": emp_id_obj, "date": att_date})
        else:
            update_data = {
                "status": status,
                "user_id": ObjectId(uid),
                "leave_reason": leave_reason or "",
                "leave_note": leave_note or ""
            }
                
            db.attendance.update_one(
                {"emp_id": emp_id_obj, "date": att_date},
                {"$set": update_data},
                upsert=True
            )
        return jsonify({'success': True, 'status': status})
    except Exception as e:
        app.logger.error("Failed to mark attendance", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/attendance/summary')
@limiter.limit("60 per minute")
@login_required
def attendance_summary():
    uid = get_current_user_id()
    month_filter = request.args.get('month', datetime.now().strftime('%Y-%m'))
    search_query = request.args.get('q', request.args.get('search', '')).strip()

    emp_query = {"user_id": ObjectId(uid)}
    if search_query:
        emp_query["$or"] = [
            {"name": {"$regex": search_query, "$options": "i"}},
            {"employee_id": {"$regex": search_query, "$options": "i"}}
        ]

    emps = list(db.employees.find(emp_query).sort("name", 1))
    emps = serialize_docs(emps)
    
    # Filter out employees who haven't joined yet by selected_date
    filtered_emps = []
    for e in emps:
        jd = e.get('joining_date')
        if jd and selected_date < jd:
            continue
        filtered_emps.append(e)
    emps = filtered_emps
    
    
    try:
        year, month_num = map(int, month_filter.split('-'))
    except ValueError:
        year, month_num = datetime.now().year, datetime.now().month
        month_filter = f"{year}-{month_num:02d}"
        
    total_days = 30
    
    current_date = datetime.now()
    if year == current_date.year and month_num == current_date.month:
        effective_days = min(current_date.day, 30)
    elif (year > current_date.year) or (year == current_date.year and month_num > current_date.month):
        effective_days = 0
    else:
        effective_days = total_days

    # Batch query for absences (Fixes N+1 issue)
    emp_ids = [ObjectId(e['id']) for e in emps] if emps else []
    absent_map = {}
    if emp_ids:
        att_cursor = db.attendance.find({
            "emp_id": {"$in": emp_ids},
            "status": "Absent",
            "date": {"$regex": f"^{month_filter}"}
        })
        absent_map = {}
        for r in att_cursor:
            wid = str(r['emp_id'])
            att_date = r['date']
            emp = next((x for x in emps if x['id'] == wid), None)
            if emp:
                jd = emp.get('joining_date')
                if jd and att_date < jd:
                    continue # Ignore absences before joining date
            absent_map[wid] = absent_map.get(wid, 0) + 1

    summary = []
    for e in emps:
        absent_days = absent_map.get(e['id'], 0)
        present_days = max(effective_days - absent_days, 0)
        summary.append({
            'id': e['id'],
            'name': e['name'],
            'employee_id': e.get('employee_id', '—'),
            'profile_image_url': e.get('profile_image_url', ''),
            'present': present_days,
            'absent': absent_days,
            'total': total_days
        })

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.args.get('ajax') == '1':
        return jsonify({
            "success": True,
            "summary": summary,
            "month_filter": month_filter,
            "search_query": search_query
        })

    return render_template('attendance_summary.html', summary=summary, month_filter=month_filter, search_query=search_query)


@app.route('/api/attendance/absent_days/<emp_id>')
@login_required
def absent_days(emp_id):
    try:
        uid = get_current_user_id()
        month_filter = request.args.get('month', datetime.now().strftime('%Y-%m'))
        
        # Verify employee belongs to current user
        emp = db.employees.find_one({"_id": ObjectId(emp_id), "user_id": ObjectId(uid)})
        if not emp:
            return jsonify({'success': False, 'message': 'Employee not found or unauthorized'}), 404

        # Fetch absent days for this month
        pipeline = [
            {"$match": {
                "emp_id": ObjectId(emp_id),
                "status": "Absent",
                "date": {"$regex": f"^{month_filter}"}
            }},
            {"$sort": {"date": 1}}
        ]
        
        absent_records = list(db.attendance.aggregate(pipeline))
        
        results = []
        for r in absent_records:
            # Parse date string "YYYY-MM-DD" to get the day name
            date_obj = datetime.strptime(r['date'], '%Y-%m-%d')
            day_name = date_obj.strftime('%A')
            # Format output date as DD-Mon-YYYY
            formatted_date = date_obj.strftime('%d-%b-%Y')
            
            results.append({
                "date": formatted_date,
                "day": day_name,
                "status": r['status'],
                "leave_reason": r.get('leave_reason', '-'),
                "leave_note": r.get('leave_note', '-')
            })

        return jsonify({'success': True, 'data': results})
    except Exception as e:
        app.logger.error("Failed to fetch absent days", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500


# ─────────────────────────────────────────
# SALARY
# ─────────────────────────────────────────

@app.route('/salary')
@limiter.limit("60 per minute")
@login_required
def salary():
    uid = get_current_user_id()
    month_filter = request.args.get('month', datetime.now().strftime('%Y-%m'))

    search = request.args.get('q', '').strip()
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 25))
    except ValueError:
        page = 1
        limit = 25
        
    skip = (page - 1) * limit
    
    query = {"user_id": ObjectId(uid)}
    if search:
        query["name"] = {"$regex": search, "$options": "i"}

    total_records = db.employees.count_documents(query)
    total_pages = (total_records + limit - 1) // limit if limit > 0 else 1

    emps = list(db.employees.find(query).sort("name", 1).skip(skip).limit(limit))
    emps = serialize_docs(emps)
    
    # Filter out employees who haven't joined yet by selected_date
    filtered_emps = []
    for e in emps:
        jd = e.get('joining_date')
        if jd and selected_date < jd:
            continue
        filtered_emps.append(e)
    emps = filtered_emps
    
    
    try:
        year, month_num = map(int, month_filter.split('-'))
    except ValueError:
        year, month_num = datetime.now().year, datetime.now().month
        month_filter = f"{year}-{month_num:02d}"
        
    total_days = 30
    
    current_date = datetime.now()
    if year == current_date.year and month_num == current_date.month:
        effective_days = min(current_date.day, 30)
    elif (year > current_date.year) or (year == current_date.year and month_num > current_date.month):
        effective_days = 0
    else:
        effective_days = total_days

    # Batch query for absences (Fixes N+1 issue)
    emp_ids = [ObjectId(e['id']) for e in emps] if emps else []
    absent_map = {}
    salary_records_map = {}
    advance_map = {}
    
    if emp_ids:
        att_cursor = db.attendance.find({
            "emp_id": {"$in": emp_ids},
            "status": "Absent",
            "date": {"$regex": f"^{month_filter}"}
        })
        absent_map = {}
        for r in att_cursor:
            wid = str(r['emp_id'])
            att_date = r['date']
            emp = next((x for x in emps if x['id'] == wid), None)
            if emp:
                jd = emp.get('joining_date')
                if jd and att_date < jd:
                    continue # Ignore absences before joining date
            absent_map[wid] = absent_map.get(wid, 0) + 1
        
        records_cursor = db.salary_records.find({"emp_id": {"$in": emp_ids}, "month": month_filter})
        for r in records_cursor:
            salary_records_map[str(r['emp_id'])] = r
            
        record_ids = [r['_id'] for r in salary_records_map.values()]
        if record_ids:
            advances_cursor = db.salary_advance_payments.find({"salary_record_id": {"$in": record_ids}})
            for adv in advances_cursor:
                wid = str(adv['salary_record_id'])
                advance_map[wid] = advance_map.get(wid, 0) + adv['amount']

    missing_emps = [e for e in emps if e['id'] not in salary_records_map]
    if missing_emps:
        docs_to_insert = []
        for e in missing_emps:
            # Calculate employee-specific effective days based on joining date
            emp_effective = effective_days
            jd_str = e.get('joining_date')
            if jd_str:
                try:
                    jy, jm, jd = map(int, jd_str.split('-'))
                    if jy == year and jm == month_num:
                        if year == current_date.year and month_num == current_date.month:
                            emp_effective = max(min(current_date.day, 30) - jd, 0)
                        else:
                            emp_effective = max(30 - jd, 0)
                    elif (year < jy) or (year == jy and month_num < jm):
                        emp_effective = 0
                except:
                    pass

            absent_days = absent_map.get(e['id'], 0)
            present_days = max(emp_effective - absent_days, 0)
            salary_per_day = e.get('salary', 0.0) / total_days
            final_salary = round(salary_per_day * present_days, 2)
            
            docs_to_insert.append({
                "_id": ObjectId(),
                "emp_id": ObjectId(e['id']),
                "month": month_filter,
                "present_days": present_days,
                "total_salary": final_salary,
                "advance_amount_paid": 0.0,
                "advance_paid_at": None,
                "payment_status": "Pending",
                "paid_at": None
            })
        if docs_to_insert:
            db.salary_records.insert_many(docs_to_insert)
            for doc in docs_to_insert:
                salary_records_map[str(doc['emp_id'])] = doc

    from pymongo import UpdateOne
    bulk_ops = []
    salary_details = []
    
    for e in emps:
        # Calculate employee-specific effective days based on joining date
        emp_effective = effective_days
        jd_str = e.get('joining_date')
        if jd_str:
            try:
                jy, jm, jd = map(int, jd_str.split('-'))
                if jy == year and jm == month_num:
                    if year == current_date.year and month_num == current_date.month:
                        emp_effective = max(min(current_date.day, 30) - jd, 0)
                    else:
                        emp_effective = max(30 - jd, 0)
                elif (year < jy) or (year == jy and month_num < jm):
                    emp_effective = 0
            except:
                pass
                
        absent_days = absent_map.get(e['id'], 0)
        present_days = max(emp_effective - absent_days, 0)
        salary_per_day = e.get('salary', 0.0) / total_days
        final_salary = round(salary_per_day * present_days, 2)
        
        rec = salary_records_map[e['id']]
        record_id_str = str(rec['_id'])
        
        update_fields = {}
        if rec.get('payment_status') in ['Unpaid', 'Pending']:
            update_fields['present_days'] = present_days
            update_fields['total_salary'] = final_salary
            
        advance_amount_paid = advance_map.get(record_id_str, 0.0)
        net_payable = round(final_salary - advance_amount_paid, 2)
        
        if advance_amount_paid > final_salary:
            payment_status = 'Overpaid'
        elif net_payable <= 0 and final_salary > 0:
            payment_status = 'Settled'
        elif final_salary == 0 and advance_amount_paid == 0:
            payment_status = 'Pending'
        else:
            payment_status = 'Pending'
            
        if rec.get('payment_status') != payment_status:
            update_fields['payment_status'] = payment_status
            
        if update_fields:
            bulk_ops.append(UpdateOne({"_id": rec['_id']}, {"$set": update_fields}))

        salary_details.append({
            'id': e['id'],
            'record_id': record_id_str,
            'name': e['name'],
            'profile_image_url': e.get('profile_image_url', ''),
            'monthly_salary': e['salary'],
            'present_days': present_days,
            'salary_per_day': round(salary_per_day, 2),
            'final_salary': final_salary,
            'advance_amount_paid': advance_amount_paid,
            'net_payable': net_payable,
            'payment_status': payment_status,
            'paid_at': rec.get('paid_at')
        })

    if bulk_ops:
        db.salary_records.bulk_write(bulk_ops)

    return render_template('salary.html', salary_details=salary_details, month_filter=month_filter, search=search, page=page, limit=limit, total_pages=total_pages, total_records=total_records)


# ─────────────────────────────────────────
# PART-TIME EMPLOYEES
# ─────────────────────────────────────────

@app.route('/part-time')
@app.route('/part-time/workers', endpoint='part_time_workers_full')
@limiter.limit("60 per minute")
@login_required
def part_time():
    show_all_workers = request.path == '/part-time/workers'
    uid = get_current_user_id()
    search = request.args.get('q', '').strip()
    selected_client = request.args.get('client', '').strip()
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 25))
    except ValueError:
        page = 1
        limit = 25
        
    skip = (page - 1) * limit
    setup_error = None
    
    pt_workers = list(db.part_time_workers.find({"user_id": ObjectId(uid)}).sort("name", 1))
    worker_info = {str(w['_id']): {"name": w['name'], "profile_image_url": w.get('profile_image_url', ''), "worker_id_str": w.get('worker_id', ''), "created_at": w.get('created_at', '')} for w in pt_workers}
    worker_ids = [ObjectId(wid) for wid in worker_info.keys()]
    
    match_query = {"worker_id": {"$in": worker_ids}}
    if selected_client:
        match_query["client_name"] = {"$regex": f"^{selected_client}$", "$options": "i"}
    if search:
        search_lower = search.lower()
        matching_workers = [
            w['_id'] for w in pt_workers 
            if search_lower in w['name'].lower() or search_lower in w.get('worker_id', '').lower()
        ]
        match_query["$or"] = [
            {"client_name": {"$regex": search, "$options": "i"}},
            {"delivery_location": {"$regex": search, "$options": "i"}},
            {"worker_id": {"$in": matching_workers}}
        ]
        
    total_records = db.part_time_work_logs.count_documents(match_query)
    total_pages = (total_records + limit - 1) // limit if limit > 0 else 1
    
    logs = list(db.part_time_work_logs.find(match_query).sort("_id", -1).skip(skip).limit(limit))
    
    # Fetch advances only for paginated logs
    log_ids = [log['_id'] for log in logs]
    advances_cursor = db.advance_payments.find({"work_log_id": {"$in": log_ids}})
    advances_map = {}
    for adv in advances_cursor:
        wid = str(adv['work_log_id'])
        advances_map[wid] = advances_map.get(wid, 0) + adv['amount']
        
    records = []
    for log in logs:
        record = serialize_doc(log)
        info = worker_info.get(str(log['worker_id']), {"name": "Unknown Worker", "profile_image_url": "", "worker_id_str": ""})
        record['worker_name'] = info['name']
        record['worker_id_str'] = info['worker_id_str']
        record['profile_image_url'] = info['profile_image_url']
        record['client_name'] = (record.get('client_name') or 'Unassigned').strip()
        record['total_price'] = float(record.get('total_price') or 0)
        record['slab_quantity'] = int(record.get('slab_quantity') or 0)
        
        total_advance = float(advances_map.get(str(log['_id']), 0))
        record['advance_paid'] = total_advance
        record['remaining_balance'] = float(record['total_price'] - total_advance)
        
        if total_advance > record['total_price']:
            record['payment_status'] = 'Overpaid'
        elif record['remaining_balance'] <= 0:
            record['payment_status'] = 'Paid'
        else:
            record['payment_status'] = 'Pending'
            
        records.append(record)

    # Lightweight analytics loop
    client_map = {}
    worker_map = {}
    
    # Pre-populate worker_map so workers with 0 jobs still appear
    for wid_obj in worker_ids:
        if search and ('$in' in match_query.get('worker_id', {})):
            if wid_obj not in match_query['worker_id']['$in'] and not (
                match_query.get('$or') and any(wid_obj in cond.get('worker_id', {}).get('$in', []) for cond in match_query['$or'] if 'worker_id' in cond)
            ):
                pass # Just let the loop add them if they match a client or location
        wid_str = str(wid_obj)
        info = worker_info.get(wid_str, {})
        worker_map[wid_str] = {
            'worker_id': wid_str,
            'worker_id_str': info.get('worker_id_str', ''),
            'name': info.get('name', 'Unknown Worker'),
            'profile_image_url': info.get('profile_image_url', ''),
            'clients': set(),
            'earnings': 0.0,
            'advances': 0.0,
            'balance': 0.0,
            'jobs': 0,
            'slabs': 0,
            'recent_assignments': [],
            'last_active_date': info.get('created_at', '')
        }
        
    # Fetch all part-time advances for the user
    all_advances = list(db.advance_payments.find({"user_id": ObjectId(uid)}, {"amount": 1, "worker_id": 1, "work_log_id": 1}))
    
    # Build log to worker mapping globally
    all_user_logs = list(db.part_time_work_logs.find({"worker_id": {"$in": worker_ids}}, {"worker_id": 1}))
    log_to_worker = {str(log['_id']): str(log['worker_id']) for log in all_user_logs}
    
    # Map each advance to its worker
    log_ids_set = {str(log['_id']) for log in all_user_logs}
    unlinked_by_worker = {}
    total_unlinked_sum = 0.0
    for adv in all_advances:
        wlid = adv.get('work_log_id')
        if not wlid or str(wlid) not in log_ids_set:
            wid_str = None
            if adv.get('worker_id'):
                wid_str = str(adv['worker_id'])
            elif wlid:
                wid_str = log_to_worker.get(str(wlid))
            if wid_str:
                unlinked_by_worker.setdefault(wid_str, []).append(adv)
                total_unlinked_sum += adv['amount']

    monthly_map = {}
    recent_clients = []
    total_payout = 0
    total_advance_all = 0
    total_remaining = 0
    total_slabs = 0

    analytics_cursor = db.part_time_work_logs.find(match_query, {
        "worker_id": 1, "client_name": 1, "total_price": 1, "slab_quantity": 1, "working_date": 1, "remaining_balance": 1
    })
    
    for log in analytics_cursor:
        client = (log.get('client_name') or 'Unassigned').strip()
        worker = worker_info.get(str(log['worker_id']), {}).get('name', 'Unknown Worker')
        t_price = float(log.get('total_price') or 0)
        r_bal = float(log.get('remaining_balance') or t_price) # fallback to t_price if missing
        adv_paid = t_price - r_bal
        sq = int(log.get('slab_quantity') or 0)
        
        total_payout += t_price
        total_advance_all += adv_paid
        total_remaining += r_bal
        total_slabs += sq

        client_stats = client_map.setdefault(client, {
            'name': client, 'entries': 0, 'total_payout': 0, 'advance_paid': 0,
            'remaining_balance': 0, 'slabs': 0, 'workers': set()
        })
        client_stats['entries'] += 1
        client_stats['total_payout'] += t_price
        client_stats['advance_paid'] += adv_paid
        client_stats['remaining_balance'] += r_bal
        client_stats['slabs'] += sq
        client_stats['workers'].add(worker)

        wid_str = str(log['worker_id'])
        worker_stats = worker_map.setdefault(wid_str, {
            'worker_id': wid_str, 
            'worker_id_str': worker_info.get(wid_str, {}).get('worker_id_str', ''),
            'name': worker, 
            'profile_image_url': worker_info.get(wid_str, {}).get('profile_image_url', ''),
            'clients': set(), 
            'earnings': 0.0, 
            'advances': 0.0,
            'balance': 0.0,
            'jobs': 0,
            'slabs': 0,
            'recent_assignments': [],
            'last_active_date': info.get('created_at', '')
        })
        worker_stats['clients'].add(client)
        worker_stats['earnings'] += t_price
        worker_stats['advances'] += adv_paid
        worker_stats['balance'] += r_bal
        worker_stats['jobs'] += 1
        worker_stats['slabs'] += sq
        working_date = log.get('working_date') or ''
        if working_date > worker_stats.get('last_active_date', ''):
            worker_stats['last_active_date'] = working_date
        if len(worker_stats['recent_assignments']) < 3:
            worker_stats['recent_assignments'].append({
                'client': client, 'date': log.get('working_date'), 'total': t_price
            })

        work_month = (log.get('working_date') or '')[:7] or 'Undated'
        monthly_map[work_month] = monthly_map.get(work_month, 0) + t_price

        if client not in recent_clients:
            recent_clients.append(client)

    # Add unlinked advances to worker stats in worker_map
    for wid_str, stats in worker_map.items():
        unlinked_sum = sum(a['amount'] for a in unlinked_by_worker.get(wid_str, []))
        stats['advances'] += unlinked_sum
        stats['balance'] = stats['earnings'] - stats['advances']

    if not selected_client:
        total_advance_all += total_unlinked_sum
        total_remaining = total_payout - total_advance_all

    client_summaries = sorted(client_map.values(), key=lambda c: c['total_payout'], reverse=True)
    for summary in client_summaries:
        summary['worker_count'] = len(summary['workers'])
        summary['workers'] = sorted(summary['workers'])
        
    paginated_records = records

    analytics = {
        'top_clients': client_summaries[:5],
        'workforce_allocation': sorted(client_summaries, key=lambda c: c['entries'], reverse=True)[:6],
        'monthly_client_expenses': [{'month': m, 'total': t} for m, t in sorted(monthly_map.items())],
        'client_productivity': sorted(client_summaries, key=lambda c: c['slabs'], reverse=True)[:6],
        'total_payout': total_payout,
        'total_advance': total_advance_all,
        'total_remaining': total_remaining,
        'total_slabs': total_slabs
    }

    worker_history = []
    # If there is a search, filter out workers with 0 jobs UNLESS they explicitly match the search term
    for worker in sorted(worker_map.values(), key=lambda w: w.get('last_active_date', ''), reverse=True):
        if search and worker['jobs'] == 0:
            search_lower = search.lower()
            if search_lower not in worker['name'].lower() and search_lower not in worker['worker_id_str'].lower():
                continue
                
        payment_status = 'Paid'
        if worker['advances'] > worker['earnings']:
            payment_status = 'Overpaid'
        elif worker['earnings'] > worker['advances']:
            payment_status = 'Pending'
                
        worker_history.append({
            'worker_id': worker['worker_id'],
            'worker_id_str': worker['worker_id_str'],
            'name': worker['name'],
            'profile_image_url': worker['profile_image_url'],
            'clients': sorted(worker['clients']),
            'earnings': worker['earnings'],
            'advances': worker['advances'],
            'balance': worker['balance'],
            'jobs': worker['jobs'],
            'slabs': worker['slabs'],
            'payment_status': payment_status,
            'recent_assignments': worker['recent_assignments']
        })

    return render_template(
        'Part_time_employee.html',
        records=paginated_records,
        total_pages=total_pages,
        current_page=page,
        total_records=total_records,
        part_time_workers=serialize_docs(pt_workers),
        setup_error=setup_error,
        search=search,
        selected_client=selected_client,
        client_summaries=client_summaries,
        recent_clients=recent_clients[:6],
        analytics=analytics,
        worker_history=worker_history,
        show_all_workers=show_all_workers
    )


@app.route('/part-time/add', methods=['POST'])
@limiter.limit("30 per minute")
@login_required
def add_part_time_work():
    uid = get_current_user_id()
    invalidate_dashboard_cache(uid)
    worker_id = request.form.get('worker_id')
    client_name = request.form.get('client_name', '').strip()
    working_date = request.form.get('working_date', '').strip()
    location = request.form.get('location', '').strip()
    
    if not worker_id or not client_name or not working_date or not location:
        return redirect(url_for('part_time'))

    try:
        slab_quantity = int(request.form.get('slab_quantity'))
        slab_price = float(request.form.get('slab_price'))
        if slab_quantity <= 0 or slab_price < 0:
            return redirect(url_for('part_time'))
    except (TypeError, ValueError):
        return redirect(url_for('part_time'))

    total_price = slab_quantity * slab_price
    worker_id_obj = safe_object_id(worker_id)
    if not worker_id_obj:
        return redirect(url_for('part_time'))

    worker = db.part_time_workers.find_one({"_id": worker_id_obj, "user_id": ObjectId(uid)})
    if not worker:
        return redirect(url_for('part_time'))

    log_doc = {
        "worker_id": worker_id_obj,
        "client_name": client_name,
        "working_date": working_date,
        "delivery_location": location,
        "slab_quantity": slab_quantity,
        "slab_price": slab_price,
        "total_price": total_price,
        "advance_paid": 0.0,
        "remaining_balance": total_price,
        "payment_status": "Pending",
        "notes": "",
        "created_at": datetime.now().isoformat()
    }
    db.part_time_work_logs.insert_one(log_doc)
    return redirect(url_for('part_time'))


def async_upload_avatar_to_cloudinary(worker_id_obj, file_bytes, filename):
    import io
    try:
        file_like = io.BytesIO(file_bytes)
        file_like.name = filename
        file_like.filename = filename
        p_url = upload_avatar_to_cloudinary(file_like)
        if p_url:
            db.part_time_workers.update_one(
                {"_id": worker_id_obj},
                {"$set": {"profile_image_url": p_url}}
            )
            app.logger.info(f"Async Cloudinary upload complete for worker {worker_id_obj}: {p_url}")
    except Exception as e:
        app.logger.error(f"Async Cloudinary upload failed for worker {worker_id_obj}: {e}")


@app.route('/part-time/workers/add', methods=['POST'])
@limiter.limit("20 per minute")
@login_required
def add_part_time_worker():
    import threading
    uid = get_current_user_id()
    name = request.form.get('name', '').strip()
    if not name:
        return redirect(url_for('part_time'))
    
    file_bytes = None
    filename = ""
    if 'profile_image' in request.files and request.files['profile_image'].filename != '':
        f = request.files['profile_image']
        filename = f.filename
        try:
            file_bytes = f.read()
        except Exception as e:
            app.logger.error(f"Failed to read upload file bytes: {e}")

    # Generate unique worker ID
    wrk_id_str = generate_worker_id()

    worker_doc = {
        "worker_id": wrk_id_str,
        "name": name,
        "profile_image_url": "",
        "user_id": ObjectId(uid),
        "created_at": datetime.now().isoformat()
    }
    result = db.part_time_workers.insert_one(worker_doc)
    worker_id_obj = result.inserted_id
    
    if file_bytes:
        threading.Thread(
            target=async_upload_avatar_to_cloudinary,
            args=(worker_id_obj, file_bytes, filename),
            daemon=True
        ).start()
    
    if request.headers.get('Accept') and 'application/json' in request.headers.get('Accept'):
        return jsonify({
            "success": True, 
            "worker_id": str(worker_id_obj),
            "worker_id_str": wrk_id_str,
            "name": name
        })
        
    return redirect(url_for('part_time'))

# ─────────────────────────────────────────
# SALARY ACTIONS
# ─────────────────────────────────────────


@app.route('/delete_part_time_worker/<worker_id>', methods=['POST'])
@login_required
def delete_part_time_worker(worker_id):
    uid = get_current_user_id()
    invalidate_dashboard_cache(uid)
    worker_id_obj = safe_object_id(worker_id)
    if not worker_id_obj:
        return redirect(url_for('part_time'))
        
    worker = db.part_time_workers.find_one({"_id": worker_id_obj, "user_id": ObjectId(uid)})
    if not worker:
        return redirect(url_for('part_time'))
        
    # Get all work logs to delete their advances
    work_logs = list(db.part_time_work_logs.find({"worker_id": worker_id_obj}))
    log_ids = [log["_id"] for log in work_logs]
    
    # Delete general advances directly linked to worker_id
    db.advance_payments.delete_many({"worker_id": worker_id_obj})
    
    if log_ids:
        db.advance_payments.delete_many({"work_log_id": {"$in": log_ids}})
        
    db.part_time_work_logs.delete_many({"worker_id": worker_id_obj})
    db.part_time_workers.delete_one({"_id": worker_id_obj, "user_id": ObjectId(uid)})
    
    return redirect(url_for('part_time'))

@app.route('/api/part-time/worker/<worker_id>/card_html', methods=['GET'])
@login_required
def get_part_time_worker_card_html(worker_id):
    from flask import render_template_string
    uid = get_current_user_id()
    try:
        worker_id_obj = ObjectId(worker_id)
    except:
        return jsonify({"success": False, "error": "Invalid worker ID"}), 400
    
    w = db.part_time_workers.find_one({"_id": worker_id_obj, "user_id": ObjectId(uid)})
    if not w:
        return jsonify({"success": False, "error": "Worker not found"}), 404
        
    w_data = {
        "worker_id": str(w["_id"]),
        "worker_id_str": w.get("worker_id", ""),
        "name": w.get("name", ""),
        "profile_image_url": w.get("profile_image_url", ""),
        "clients": [],
        "earnings": 0,
        "advances": 0,
        "balance": 0,
        "jobs": 0,
        "slabs": 0
    }
    
    card_template = """
    <div class="dashboard-worker-card bg-white dark:bg-[#1a222c] glass-panel flex flex-col rounded-2xl overflow-hidden transition-all hover:shadow-lg border border-slate-200 dark:border-white/10" data-name="{{ w['name']|lower }}" data-id="{{ w['worker_id_str']|lower }}" data-clients="{{ w['clients']|join(' ')|lower }}">
      <div class="p-5 flex items-start justify-between border-b border-slate-100 dark:border-white/5 relative">
        <div class="flex items-center gap-3 cursor-pointer group" onclick="openWorkerProfile('{{ w['worker_id'] }}', '{{ w['name']|escape }}', '{{ w['worker_id_str']|escape }}', '{{ w['profile_image_url']|escape }}')">
          {% if w['profile_image_url'] %}
            <img src="{{ w['profile_image_url'] }}" class="w-12 h-12 rounded-xl object-cover shadow-sm group-hover:scale-105 transition-transform border border-slate-200 dark:border-white/10">
          {% else %}
            <div class="w-12 h-12 rounded-xl bg-gradient-to-br from-primary to-purpleaccent text-white flex items-center justify-center text-lg font-bold shadow-sm group-hover:scale-105 transition-transform border border-white dark:border-white/10">
              {{ w['name'][:1]|upper }}
            </div>
          {% endif %}
          <div class="flex flex-col">
            <h3 class="font-bold text-slate-900 dark:text-white leading-tight group-hover:text-primary transition-colors text-base">{{ w['name'] }}</h3>
            <span class="text-xs font-mono text-slate-500 dark:text-slate-400 font-medium uppercase tracking-wider">Part-Time</span>
          </div>
        </div>
        <button class="w-8 h-8 rounded-full flex items-center justify-center text-slate-400 hover:text-primary hover:bg-slate-100 dark:hover:bg-white/10 transition-all" onclick="openWorkerProfile('{{ w['worker_id'] }}', '{{ w['name']|escape }}', '{{ w['worker_id_str']|escape }}', '{{ w['profile_image_url']|escape }}')">
          <i class="bi bi-three-dots-vertical"></i>
        </button>
      </div>

      <div class="p-5 grid grid-cols-3 gap-4 border-b border-slate-100 dark:border-white/5">
        <div class="flex flex-col">
          <span class="text-[10px] text-slate-500 dark:text-slate-400 uppercase tracking-widest font-bold mb-1">Earnings</span>
          <span class="font-mono text-sm font-bold text-slate-900 dark:text-white" id="card-earnings-{{ w['worker_id'] }}">₹0</span>
        </div>
        <div class="flex flex-col">
          <span class="text-[10px] text-slate-500 dark:text-slate-400 uppercase tracking-widest font-bold mb-1">Advances</span>
          <span class="font-mono text-sm font-bold text-danger" id="card-advances-{{ w['worker_id'] }}">₹0</span>
        </div>
        <div class="flex flex-col text-right">
          <span class="text-[10px] text-slate-500 dark:text-slate-400 uppercase tracking-widest font-bold mb-1">Balance</span>
          <span id="card-balance-{{ w['worker_id'] }}" class="font-mono text-[15px] font-black text-slate-900 dark:text-white">₹0</span>
        </div>
      </div>

      <div class="p-4 flex items-center justify-between bg-slate-50/50 dark:bg-white/[0.02]">
        <div class="flex items-center gap-1.5 text-xs font-medium text-slate-600 dark:text-slate-300">
          <i class="bi bi-briefcase text-slate-400"></i>
          <span><span id="card-jobs-{{ w['worker_id'] }}">0</span> Jobs (<span id="card-slabs-{{ w['worker_id'] }}">0</span> Slabs)</span>
        </div>
        <div>
          <span id="card-status-{{ w['worker_id'] }}" class="px-2 py-1 rounded-md text-[10px] font-bold uppercase tracking-wider bg-slate-100 text-slate-500 dark:bg-white/5 dark:text-slate-400">No Work</span>
        </div>
      </div>

      <div class="p-4 grid grid-cols-3 gap-2 border-t border-slate-100 dark:border-white/5 bg-slate-50/50 dark:bg-white/[0.02]">
        <button data-action="log-work" data-worker-id="{{ w['worker_id'] }}" data-worker-name="{{ w['name']|escape }}" class="flex-1 min-h-[40px] btn-saas btn-saas-sm bg-primary hover:bg-primary/90 text-white shadow-sm font-semibold text-[11px] rounded-xl transition-colors">
          <i class="bi bi-file-earmark-plus mr-1"></i> Log Work
        </button>
        <button data-action="advance" data-worker-id="{{ w['worker_id'] }}" data-worker-name="{{ w['name']|escape }}" data-worker-id-str="{{ w['worker_id_str']|escape }}" class="flex-1 min-h-[40px] btn-saas btn-saas-sm bg-slate-100 hover:bg-slate-200 text-slate-800 dark:bg-white/5 dark:hover:bg-white/10 dark:text-slate-200 font-semibold text-[11px] rounded-xl transition-colors">
          <i class="bi bi-cash-stack mr-1"></i> Advance
        </button>
        <button data-action="ledger" data-worker-id="{{ w['worker_id'] }}" data-worker-name="{{ w['name']|escape }}" class="flex-1 min-h-[40px] btn-saas btn-saas-sm bg-slate-100 hover:bg-slate-200 text-slate-800 dark:bg-white/5 dark:hover:bg-white/10 dark:text-slate-200 font-semibold text-[11px] rounded-xl transition-colors">
          <i class="bi bi-journal-text mr-1"></i> Ledger
        </button>
      </div>
      
      <div id="ledger-container-{{ w['worker_id'] }}" class="hidden border-t border-slate-200 dark:border-white/10 bg-slate-50 dark:bg-[#151b23]">
        <div class="p-4 flex items-center justify-between border-b border-slate-200 dark:border-white/5">
          <h4 class="font-bold text-sm text-slate-900 dark:text-white flex items-center gap-2">
            <i class="bi bi-journal-text text-primary"></i> Ledger
          </h4>
          <button onclick="toggleWorkerLedger('{{ w['worker_id'] }}')" class="w-6 h-6 flex items-center justify-center rounded-full bg-slate-200 hover:bg-slate-300 dark:bg-white/10 dark:hover:bg-white/20 text-slate-500 transition-colors">
            <i class="bi bi-x-lg text-[10px]"></i>
          </button>
        </div>
        <div class="p-4 overflow-x-auto custom-scrollbar">
          <div id="ledger-content-{{ w['worker_id'] }}" class="min-w-[800px]">
            <div class="py-10 text-center text-slate-500">
              <div class="spinner-border text-primary inline-block w-6 h-6 border-2 rounded-full animate-spin border-t-transparent"></div>
              <p class="mt-2 text-sm font-medium">Loading ledger...</p>
            </div>
          </div>
        </div>
      </div>
    </div>
    """
    
    html = render_template_string(card_template, w=w_data)
    return jsonify({"success": True, "html": html})

@app.route('/salary/mark_paid', methods=['POST'])
@limiter.limit("30 per minute")         # prevent accidental bulk payment triggers
@login_required
def mark_paid():
    data = request.get_json() or {}
    emp_id = data.get('emp_id')
    month = data.get('month')
    paid_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    emp_id_obj = safe_object_id(emp_id)
    if not emp_id_obj:
        return jsonify({'success': False, 'message': 'Invalid Employee ID'}), 400

    uid = get_current_user_id()
    # Verify employee belongs to current user
    emp = db.employees.find_one({"_id": emp_id_obj, "user_id": ObjectId(uid)})
    if not emp:
        return jsonify({'success': False, 'message': 'Unauthorized or Employee not found'}), 403

    # Find the corresponding salary record
    rec = db.salary_records.find_one({"emp_id": emp_id_obj, "month": month})
    if not rec:
        return jsonify({'success': False, 'message': 'Salary record not found'}), 404

    # Calculate current advances
    advances = list(db.salary_advance_payments.find({"salary_record_id": rec['_id']}))
    total_advances = sum(a['amount'] for a in advances)
    total_salary = float(rec.get('total_salary', 0))
    net_payable = round(total_salary - total_advances, 2)

    # Prevent double payment
    if rec.get('payment_status') == 'Settled':
        return jsonify({'success': False, 'message': 'Salary already settled'}), 400

    # If there's a remaining balance, record a final settlement advance payment
    if net_payable > 0:
        db.salary_advance_payments.insert_one({
            "salary_record_id": rec['_id'],
            "emp_id": emp_id_obj,
            "user_id": ObjectId(uid),
            "month": month,
            "amount": net_payable,
            "payment_date": datetime.now().strftime('%Y-%m-%d'),
            "notes": "Final Salary Settlement",
            "is_migrated": False,
            "created_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat()
        })
        total_advances += net_payable
        net_payable = 0.0

    db.salary_records.update_one(
        {"_id": rec['_id']},
        {"$set": {
            "payment_status": "Settled",
            "paid_at": paid_at,
            "advance_amount_paid": total_advances
        }}
    )

    return jsonify({
        'success': True,
        'total_advance': total_advances,
        'net_payable': net_payable,
        'payment_status': 'Settled',
        'paid_at': paid_at
    })


@app.route('/salary/set_advance', methods=['POST'])
@limiter.limit("30 per minute")
@login_required
def salary_set_advance():
    data = request.get_json() or {}
    emp_id = data.get('emp_id')
    month = data.get('month')
    advance_amount_paid = data.get('advance_amount_paid')

    emp_id_obj = safe_object_id(emp_id)
    if not emp_id_obj:
        return jsonify({'success': False, 'message': 'Invalid Employee ID'}), 400

    uid = get_current_user_id()
    # Verify employee belongs to current user
    emp = db.employees.find_one({"_id": emp_id_obj, "user_id": ObjectId(uid)})
    if not emp:
        return jsonify({'success': False, 'message': 'Unauthorized or Employee not found'}), 403

    try:
        advance_amount_paid = float(advance_amount_paid)
        if advance_amount_paid < 0:
            return jsonify({'success': False, 'message': 'Advance must be >= 0'}), 400
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'Advance must be a number'}), 400

    advance_paid_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    db.salary_records.update_one(
        {"emp_id": emp_id_obj, "month": month, "payment_status": {"$in": ["Unpaid", "Paid"]}},
        {"$set": {
            "advance_amount_paid": advance_amount_paid,
            "advance_paid_at": advance_paid_at
        }}
    )

    # Recalculate net payable from current salary record
    rec = db.salary_records.find_one({"emp_id": emp_id_obj, "month": month})
    earned = rec['total_salary'] if rec else 0
    net_payable = round(earned - advance_amount_paid, 2)

    return jsonify({
        'success': True,
        'advance_amount_paid': advance_amount_paid,
        'advance_paid_at': advance_paid_at,
        'net_payable': net_payable
    })


# ─────────────────────────────────────────
# SALARY ADVANCE LEDGER APIs
# ─────────────────────────────────────────

@app.route('/api/salary/advance/add', methods=['POST'])
@limiter.limit("30 per minute")
@login_required
def api_salary_advance_add():
    uid = get_current_user_id()
    data = request.get_json() or {}
    record_id = data.get('record_id')
    amount = data.get('amount')
    payment_date = data.get('payment_date')
    notes = data.get('notes', '')

    try:
        amount = float(amount)
        if amount <= 0:
            return jsonify({'success': False, 'message': 'Amount must be > 0'}), 400
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'Invalid amount'}), 400

    record_id_obj = safe_object_id(record_id)
    if not record_id_obj:
        return jsonify({'success': False, 'message': 'Invalid Record ID'}), 400

    rec = db.salary_records.find_one({"_id": record_id_obj})
    if not rec:
        return jsonify({'success': False, 'message': 'Salary record not found'}), 404

    emp = db.employees.find_one({"_id": rec['emp_id'], "user_id": ObjectId(uid)})
    if not emp:
        return jsonify({'success': False, 'message': 'Not authorized'}), 403

    db.salary_advance_payments.insert_one({
        "salary_record_id": record_id_obj,
        "emp_id": rec['emp_id'],
        "user_id": ObjectId(uid),
        "month": rec.get('month'),
        "amount": amount,
        "payment_date": payment_date,
        "notes": notes,
        "is_migrated": False,
        "created_at": datetime.now().isoformat(),
        "updated_at": datetime.now().isoformat()
    })

    advances = list(db.salary_advance_payments.find({"salary_record_id": record_id_obj}))
    total_adv = sum(a['amount'] for a in advances)
    total_salary = float(rec.get('total_salary', 0))
    net_payable = round(total_salary - total_adv, 2)

    if total_adv > total_salary:
        status = 'Overpaid'
    elif net_payable <= 0 and total_salary > 0:
        status = 'Settled'
    else:
        status = 'Pending'

    db.salary_records.update_one({"_id": record_id_obj}, {"$set": {"payment_status": status}})

    return jsonify({'success': True, 'total_advance': total_adv, 'net_payable': net_payable, 'payment_status': status})


@app.route('/api/salary/advance/edit', methods=['POST'])
@limiter.limit("30 per minute")
@login_required
def api_salary_advance_edit():
    uid = get_current_user_id()
    data = request.get_json() or {}
    adv_id = data.get('advance_id')
    amount = data.get('amount')
    payment_date = data.get('payment_date')
    notes = data.get('notes', '')

    try:
        amount = float(amount)
        if amount <= 0:
            return jsonify({'success': False, 'message': 'Amount must be > 0'}), 400
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'Invalid amount'}), 400

    adv_id_obj = safe_object_id(adv_id)
    adv = db.salary_advance_payments.find_one({"_id": adv_id_obj})
    if not adv:
        return jsonify({'success': False, 'message': 'Advance not found'}), 404

    emp = db.employees.find_one({"_id": adv['emp_id'], "user_id": ObjectId(uid)})
    if not emp:
        return jsonify({'success': False, 'message': 'Not authorized'}), 403

    db.salary_advance_payments.update_one({"_id": adv_id_obj}, {"$set": {
        "amount": amount, "payment_date": payment_date,
        "notes": notes, "updated_at": datetime.now().isoformat()
    }})

    rec = db.salary_records.find_one({"_id": adv['salary_record_id']})
    advances = list(db.salary_advance_payments.find({"salary_record_id": adv['salary_record_id']}))
    total_adv = sum(a['amount'] for a in advances)
    total_salary = float(rec.get('total_salary', 0)) if rec else 0
    net_payable = round(total_salary - total_adv, 2)

    if total_adv > total_salary:
        status = 'Overpaid'
    elif net_payable <= 0 and total_salary > 0:
        status = 'Settled'
    else:
        status = 'Pending'

    db.salary_records.update_one({"_id": adv['salary_record_id']}, {"$set": {"payment_status": status}})

    return jsonify({'success': True, 'total_advance': total_adv, 'net_payable': net_payable, 'payment_status': status})


@app.route('/api/salary/advance/delete', methods=['POST'])
@limiter.limit("30 per minute")
@login_required
def api_salary_advance_delete():
    uid = get_current_user_id()
    data = request.get_json() or {}
    adv_id = data.get('advance_id')

    adv_id_obj = safe_object_id(adv_id)
    adv = db.salary_advance_payments.find_one({"_id": adv_id_obj})
    if not adv:
        return jsonify({'success': False, 'message': 'Advance not found'}), 404

    emp = db.employees.find_one({"_id": adv['emp_id'], "user_id": ObjectId(uid)})
    if not emp:
        return jsonify({'success': False, 'message': 'Not authorized'}), 403

    db.salary_advance_payments.delete_one({"_id": adv_id_obj})

    rec = db.salary_records.find_one({"_id": adv['salary_record_id']})
    advances = list(db.salary_advance_payments.find({"salary_record_id": adv['salary_record_id']}))
    total_adv = sum(a['amount'] for a in advances)
    total_salary = float(rec.get('total_salary', 0)) if rec else 0
    net_payable = round(total_salary - total_adv, 2)

    if total_adv > total_salary:
        status = 'Overpaid'
    elif net_payable <= 0 and total_salary > 0:
        status = 'Settled'
    else:
        status = 'Pending'

    db.salary_records.update_one({"_id": adv['salary_record_id']}, {"$set": {"payment_status": status}})

    return jsonify({'success': True, 'total_advance': total_adv, 'net_payable': net_payable, 'payment_status': status})


@app.route('/api/salary/advance/history/<record_id>', methods=['GET'])
@login_required
def api_salary_advance_history(record_id):
    uid = get_current_user_id()
    record_id_obj = safe_object_id(record_id)

    rec = db.salary_records.find_one({"_id": record_id_obj})
    if not rec:
        return jsonify({'success': False, 'message': 'Not found'}), 404

    emp = db.employees.find_one({"_id": rec['emp_id'], "user_id": ObjectId(uid)})
    if not emp:
        return jsonify({'success': False, 'message': 'Not authorized'}), 403

    advances = list(db.salary_advance_payments.find({"salary_record_id": record_id_obj}).sort("created_at", 1))
    return jsonify({'success': True, 'advances': serialize_docs(advances)})


@app.route('/api/part-time/search', methods=['GET'])
@login_required
def api_part_time_search():
    uid = get_current_user_id()
    search = request.args.get('q', '').strip()
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 25))
    except ValueError:
        page = 1
        limit = 25
    
    # Get decoupled part-time workers map
    pt_workers = list(db.part_time_workers.find({"user_id": ObjectId(uid)}))
    worker_names = {w['_id']: w['name'] for w in pt_workers}
    all_worker_ids = list(worker_names.keys())
    
    query = {"worker_id": {"$in": all_worker_ids}}
    
    if search:
        import re
        escaped_search = re.escape(search)
        matching_worker_ids = [w_id for w_id, w_name in worker_names.items() if re.search(escaped_search, w_name, re.IGNORECASE)]
        
        query["$and"] = [
            {"worker_id": {"$in": all_worker_ids}},
            {"$or": [
                {"worker_id": {"$in": matching_worker_ids}},
                {"client_name": {"$regex": escaped_search, "$options": "i"}},
                {"delivery_location": {"$regex": escaped_search, "$options": "i"}}
            ]}
        ]
        
    total_records = db.part_time_work_logs.count_documents(query)
    total_pages = max((total_records + limit - 1) // limit, 1) if limit > 0 else 1
    
    logs = list(db.part_time_work_logs.find(query).sort("_id", -1).skip((page-1)*limit).limit(limit))
    
    log_ids = [log['_id'] for log in logs]
    advances_cursor = db.advance_payments.find({"work_log_id": {"$in": log_ids}})
    advances_map = {}
    for adv in advances_cursor:
        wid = str(adv['work_log_id'])
        advances_map[wid] = advances_map.get(wid, 0) + adv['amount']
        
    results = []
    for log in logs:
        wname = worker_names.get(log['worker_id'], 'Unknown')
        cname = log.get('client_name', '')
        
        record = serialize_doc(log)
        record['worker_name'] = wname
        record['client_name'] = cname
        record['total_price'] = float(record.get('total_price') or 0)
        record['slab_quantity'] = int(record.get('slab_quantity') or 0)
        
        total_advance = float(advances_map.get(str(log['_id']), 0))
        record['advance_paid'] = total_advance
        record['remaining_balance'] = float(record['total_price'] - total_advance)
        
        if total_advance > record['total_price']:
            record['payment_status'] = 'Overpaid'
        elif record['remaining_balance'] <= 0:
            record['payment_status'] = 'Paid'
        else:
            record['payment_status'] = 'Pending'
            
        results.append(record)
        
    return jsonify({
        'success': True,
        'records': results,
        'total_pages': total_pages,
        'current_page': page,
        'total_records': total_records
    })



@app.route('/api/part-time/advance/add', methods=['POST'])
@login_required
def api_add_advance():
    data = request.get_json() or {}
    record_id = data.get('record_id')
    worker_id = data.get('worker_id')
    amount = data.get('amount')
    payment_date = data.get('payment_date')
    notes = data.get('notes', '')
    
    try:
        amount = float(amount)
        if amount <= 0:
            return jsonify({'success': False, 'message': 'Amount must be > 0'}), 400
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'Invalid amount'}), 400
        
    uid = get_current_user_id()
    
    # 1. If record_id is provided, it's a linked advance
    if record_id:
        record_id_obj = safe_object_id(record_id)
        if not record_id_obj:
            return jsonify({'success': False, 'message': 'Invalid ID'}), 400
            
        log_record = db.part_time_work_logs.find_one({"_id": record_id_obj})
        if not log_record:
            return jsonify({'success': False, 'message': 'Not found'}), 404
            
        worker = db.part_time_workers.find_one({"_id": log_record['worker_id'], "user_id": ObjectId(uid)})
        if not worker:
            return jsonify({'success': False, 'message': 'Not found'}), 404
            
        worker_id_obj = log_record['worker_id']
    else:
        # 2. Unlinked advance (General Advance)
        if not worker_id:
            return jsonify({'success': False, 'message': 'Worker ID is required for general advance'}), 400
        worker_id_obj = safe_object_id(worker_id)
        if not worker_id_obj:
            return jsonify({'success': False, 'message': 'Invalid Worker ID'}), 400
            
        worker = db.part_time_workers.find_one({"_id": worker_id_obj, "user_id": ObjectId(uid)})
        if not worker:
            return jsonify({'success': False, 'message': 'Worker not found'}), 404
            
        record_id_obj = None
        
    db.advance_payments.insert_one({
        "work_log_id": record_id_obj,
        "worker_id": worker_id_obj,
        "user_id": ObjectId(uid),
        "amount": amount,
        "payment_date": payment_date,
        "notes": notes,
        "created_at": datetime.now().isoformat(),
        "updated_at": datetime.now().isoformat()
    })
    
    db.audit_logs.insert_one({
        "user_id": ObjectId(uid),
        "module": "Part-Time Advance",
        "action": "ADD",
        "details": f"Added advance of ₹{amount} for worker {str(worker_id_obj)}" + (f" (work log {str(record_id_obj)})" if record_id_obj else " (General Advance)"),
        "timestamp": datetime.now().isoformat()
    })
    
    if record_id_obj:
        advances = list(db.advance_payments.find({"work_log_id": record_id_obj}))
        total_adv = sum(a['amount'] for a in advances)
        total_price = float(log_record['total_price'] or 0)
        rem_bal = total_price - total_adv
        
        status = 'Pending'
        if total_adv > total_price: status = 'Overpaid'
        elif rem_bal <= 0: status = 'Paid'
        
        # Update DB for backward compat/analytics caching
        db.part_time_work_logs.update_one({"_id": record_id_obj}, {"$set": {"payment_status": status, "remaining_balance": rem_bal}})
        
        return jsonify({
            'success': True,
            'total_advance': total_adv,
            'remaining_balance': rem_bal,
            'payment_status': status
        })
    else:
        return jsonify({
            'success': True,
            'total_advance': amount,
            'remaining_balance': 0.0,
            'payment_status': 'Paid'
        })


@app.route('/api/part-time/advance/edit', methods=['POST'])
@login_required
def api_edit_advance():
    data = request.get_json() or {}
    adv_id = data.get('advance_id')
    amount = data.get('amount')
    payment_date = data.get('payment_date')
    notes = data.get('notes', '')
    
    try:
        amount = float(amount)
        if amount <= 0: return jsonify({'success': False, 'message': 'Amount must be > 0'}), 400
    except:
        return jsonify({'success': False, 'message': 'Invalid amount'}), 400
        
    adv_id_obj = safe_object_id(adv_id)
    adv_record = db.advance_payments.find_one({"_id": adv_id_obj})
    if not adv_record: return jsonify({'success': False, 'message': 'Advance not found'}), 404
    
    uid = get_current_user_id()
    
    # Authorize using worker_id directly (fallback to work_log_id for older records)
    worker_id = adv_record.get('worker_id')
    if not worker_id and adv_record.get('work_log_id'):
        log_record = db.part_time_work_logs.find_one({"_id": adv_record['work_log_id']})
        if log_record:
            worker_id = log_record.get('worker_id')
            
    if not worker_id:
        return jsonify({'success': False, 'message': 'Not authorized'}), 403
        
    worker = db.part_time_workers.find_one({"_id": ObjectId(worker_id), "user_id": ObjectId(uid)})
    if not worker: return jsonify({'success': False, 'message': 'Not authorized'}), 403
    
    db.advance_payments.update_one({"_id": adv_id_obj}, {"$set": {
        "amount": amount, "payment_date": payment_date, "notes": notes, "updated_at": datetime.now().isoformat()
    }})
    
    db.audit_logs.insert_one({
        "user_id": ObjectId(uid),
        "module": "Part-Time Advance",
        "action": "EDIT",
        "details": f"Edited advance {str(adv_id_obj)} to ₹{amount}",
        "timestamp": datetime.now().isoformat()
    })
    
    # recalculate if linked
    work_log_id = adv_record.get('work_log_id')
    if work_log_id:
        log_record = db.part_time_work_logs.find_one({"_id": work_log_id})
        if log_record:
            advances = list(db.advance_payments.find({"work_log_id": work_log_id}))
            total_adv = sum(a['amount'] for a in advances)
            total_price = float(log_record['total_price'] or 0)
            rem_bal = total_price - total_adv
            
            status = 'Pending'
            if total_adv > total_price: status = 'Overpaid'
            elif rem_bal <= 0: status = 'Paid'
            
            db.part_time_work_logs.update_one({"_id": work_log_id}, {"$set": {"payment_status": status, "remaining_balance": rem_bal}})
            
            return jsonify({
                'success': True,
                'total_advance': total_adv,
                'remaining_balance': rem_bal,
                'payment_status': status
            })
            
    return jsonify({
        'success': True,
        'total_advance': amount,
        'remaining_balance': 0.0,
        'payment_status': 'Paid'
    })


@app.route('/api/part-time/advance/delete', methods=['POST'])
@login_required
def api_delete_advance():
    data = request.get_json() or {}
    adv_id = data.get('advance_id')
    adv_id_obj = safe_object_id(adv_id)
    
    adv_record = db.advance_payments.find_one({"_id": adv_id_obj})
    if not adv_record: return jsonify({'success': False, 'message': 'Advance not found'}), 404
    
    uid = get_current_user_id()
    
    # Authorize using worker_id directly (fallback to work_log_id for older records)
    worker_id = adv_record.get('worker_id')
    if not worker_id and adv_record.get('work_log_id'):
        log_record = db.part_time_work_logs.find_one({"_id": adv_record['work_log_id']})
        if log_record:
            worker_id = log_record.get('worker_id')
            
    if not worker_id:
        return jsonify({'success': False, 'message': 'Not authorized'}), 403
        
    worker = db.part_time_workers.find_one({"_id": ObjectId(worker_id), "user_id": ObjectId(uid)})
    if not worker: return jsonify({'success': False, 'message': 'Not authorized'}), 403
    
    db.advance_payments.delete_one({"_id": adv_id_obj})
    
    db.audit_logs.insert_one({
        "user_id": ObjectId(uid),
        "module": "Part-Time Advance",
        "action": "DELETE",
        "details": f"Deleted advance {str(adv_id_obj)} of ₹{adv_record['amount']}",
        "timestamp": datetime.now().isoformat()
    })
    
    # recalculate if linked
    work_log_id = adv_record.get('work_log_id')
    if work_log_id:
        log_record = db.part_time_work_logs.find_one({"_id": work_log_id})
        if log_record:
            advances = list(db.advance_payments.find({"work_log_id": work_log_id}))
            total_adv = sum(a['amount'] for a in advances)
            total_price = float(log_record['total_price'] or 0)
            rem_bal = total_price - total_adv
            
            status = 'Pending'
            if total_adv > total_price: status = 'Overpaid'
            elif rem_bal <= 0: status = 'Paid'
            
            db.part_time_work_logs.update_one({"_id": work_log_id}, {"$set": {"payment_status": status, "remaining_balance": rem_bal}})
            
            return jsonify({
                'success': True,
                'total_advance': total_adv,
                'remaining_balance': rem_bal,
                'payment_status': status
            })
            
    return jsonify({
        'success': True,
        'total_advance': 0.0,
        'remaining_balance': 0.0,
        'payment_status': 'Paid'
    })


@app.route('/api/part-time/mark_paid', methods=['POST'])
@login_required
def api_part_time_mark_paid():
    data = request.get_json() or {}
    record_id = data.get('record_id')
    uid = get_current_user_id()
    
    record_id_obj = safe_object_id(record_id)
    if not record_id_obj:
        return jsonify({'success': False, 'message': 'Invalid ID'}), 400
        
    log_record = db.part_time_work_logs.find_one({"_id": record_id_obj})
    if not log_record:
        return jsonify({'success': False, 'message': 'Not found'}), 404
        
    worker = db.part_time_workers.find_one({"_id": log_record['worker_id'], "user_id": ObjectId(uid)})
    if not worker:
        return jsonify({'success': False, 'message': 'Not found'}), 404
        
    advances = list(db.advance_payments.find({"work_log_id": record_id_obj}))
    total_adv = sum(a['amount'] for a in advances)
    total_price = float(log_record['total_price'] or 0)
    rem_bal = total_price - total_adv
    
    if rem_bal > 0:
        amount = round(rem_bal, 2)
        payment_date = datetime.now().strftime('%Y-%m-%d')
        notes = "Auto-settled via Mark Paid"
        
        db.advance_payments.insert_one({
            "work_log_id": record_id_obj,
            "user_id": ObjectId(uid),
            "amount": amount,
            "payment_date": payment_date,
            "notes": notes,
            "created_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat()
        })
        
        db.audit_logs.insert_one({
            "user_id": ObjectId(uid),
            "module": "Part-Time Advance",
            "action": "MARK_PAID",
            "details": f"Auto-settled balance of ₹{amount} for work log {str(record_id_obj)}",
            "timestamp": datetime.now().isoformat()
        })
        
        # recalculate
        total_adv += amount
        rem_bal = total_price - total_adv
        
    status = 'Pending'
    if total_adv > total_price: status = 'Overpaid'
    elif rem_bal <= 0: status = 'Paid'
    
    db.part_time_work_logs.update_one({"_id": record_id_obj}, {"$set": {"payment_status": status, "remaining_balance": rem_bal}})
    
    return jsonify({
        'success': True,
        'total_advance': total_adv,
        'remaining_balance': rem_bal,
        'payment_status': status
    })



@app.route('/api/part-time/advance/history/<work_log_id>', methods=['GET'])
@login_required
def api_advance_history(work_log_id):
    uid = get_current_user_id()
    w_id_obj = safe_object_id(work_log_id)
    
    log_record = db.part_time_work_logs.find_one({"_id": w_id_obj})
    if not log_record: return jsonify({'success': False, 'message': 'Not found'}), 404
    worker = db.part_time_workers.find_one({"_id": log_record['worker_id'], "user_id": ObjectId(uid)})
    if not worker: return jsonify({'success': False, 'message': 'Not authorized'}), 403
    
    advances = list(db.advance_payments.find({"work_log_id": w_id_obj}).sort("created_at", 1))
    return jsonify({'success': True, 'advances': serialize_docs(advances)})


@app.route('/api/part-time/worker/<worker_id>/summary', methods=['GET'])
@login_required
def api_worker_summary(worker_id):
    uid = get_current_user_id()
    w_id_obj = safe_object_id(worker_id)
    
    worker = db.part_time_workers.find_one({"_id": w_id_obj, "user_id": ObjectId(uid)})
    if not worker: return jsonify({'success': False, 'message': 'Not found'}), 404
    
    logs = list(db.part_time_work_logs.find({"worker_id": w_id_obj}))
    log_ids = [l['_id'] for l in logs]
    
    advances_cursor = db.advance_payments.find({
        "$or": [
            {"work_log_id": {"$in": log_ids}},
            {"worker_id": w_id_obj}
        ]
    })
    # Deduplicate by _id
    adv_dict = {str(a['_id']): a for a in advances_cursor}
    total_advance = sum(a['amount'] for a in adv_dict.values())
    
    total_jobs = len(logs)
    total_slabs = sum(l.get('slab_quantity', 0) for l in logs)
    total_earnings = sum(l.get('total_price', 0) for l in logs)
    outstanding_balance = total_earnings - total_advance
    
    return jsonify({
        'success': True,
        'name': worker.get('name', 'Unknown'),
        'worker_id_str': worker.get('worker_id', '—'),
        'total_jobs': total_jobs,
        'total_slabs': total_slabs,
        'total_earnings': total_earnings,
        'total_advance': total_advance,
        'outstanding_balance': outstanding_balance
    })

@app.route('/api/part-time/worker/<worker_id>/ledger', methods=['GET'])
@login_required
def api_worker_ledger(worker_id):
    uid = get_current_user_id()
    w_id_obj = safe_object_id(worker_id)
    
    worker = db.part_time_workers.find_one({"_id": w_id_obj, "user_id": ObjectId(uid)})
    if not worker: return jsonify({'success': False, 'message': 'Not found'}), 404
    
    logs = list(db.part_time_work_logs.find({"worker_id": w_id_obj}).sort("_id", -1))
    log_ids = [l['_id'] for l in logs]
    
    # Query all advances for this worker
    all_advances = list(db.advance_payments.find({
        "$or": [
            {"work_log_id": {"$in": log_ids}},
            {"worker_id": w_id_obj}
        ]
    }))
    
    # Map work log ID to its linked advances
    log_ids_strs = {str(lid) for lid in log_ids}
    advances_map = {}
    unlinked_advances = []
    
    for adv in all_advances:
        wlid = adv.get('work_log_id')
        if wlid and str(wlid) in log_ids_strs:
            wid = str(wlid)
            advances_map[wid] = advances_map.get(wid, 0) + adv['amount']
        else:
            unlinked_advances.append(adv)
        
    records = []
    for log in logs:
        record = serialize_doc(log)
        record['worker_name'] = worker.get('name', 'Unknown')
        record['client_name'] = (record.get('client_name') or 'Unassigned').strip()
        record['total_price'] = float(record.get('total_price') or 0)
        record['slab_quantity'] = int(record.get('slab_quantity') or 0)
        
        total_advance = float(advances_map.get(str(log['_id']), 0))
        record['advance_paid'] = total_advance
        record['remaining_balance'] = float(record['total_price'] - total_advance)
        
        if total_advance > record['total_price']:
            record['payment_status'] = 'Overpaid'
        elif record['remaining_balance'] <= 0:
            record['payment_status'] = 'Paid'
        else:
            record['payment_status'] = 'Pending'
            
        records.append(record)
        
    for adv in unlinked_advances:
        record = {
            'id': str(adv['_id']),
            'is_advance': True,
            'working_date': adv.get('payment_date') or adv.get('created_at', '')[:10],
            'client_name': 'General Salary Advance',
            'delivery_location': adv.get('notes', '').strip() or '—',
            'slab_quantity': 0,
            'slab_price': 0.0,
            'total_price': 0.0,
            'advance_paid': float(adv.get('amount') or 0),
            'remaining_balance': -float(adv.get('amount') or 0),
            'payment_status': 'Paid',
            'worker_name': worker.get('name', 'Unknown'),
            'notes': adv.get('notes', '')
        }
        records.append(record)
        
    # Sort combined records by date descending
    records.sort(key=lambda r: r.get('working_date', ''), reverse=True)
    
    return jsonify({
        'success': True,
        'records': records
    })

# ─────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────

# ─────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────

@app.route('/api/export/options')
@limiter.limit("60 per minute")
@login_required
def get_export_options():
    uid = get_current_user_id()
    employees = list(db.employees.find({"user_id": ObjectId(uid)}).sort("name", 1))
    pt_workers = list(db.part_time_workers.find({"user_id": ObjectId(uid)}).sort("name", 1))
    
    return jsonify({
        "success": True,
        "employees": [{"id": str(e['_id']), "name": e.get('name')} for e in employees],
        "part_time_workers": [{"id": str(w['_id']), "name": w.get('name')} for w in pt_workers]
    })

# ─────────────────────────────────────────
# AVATAR UPLOAD (Cloudinary)
# ─────────────────────────────────────────
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def upload_avatar_to_cloudinary(file):
    if not file or file.filename == '' or not allowed_file(file.filename):
        return None
    file.seek(0, os.SEEK_END)
    if file.tell() > 5 * 1024 * 1024:
        return None
    file.seek(0)
    try:
        res = cloudinary.uploader.upload(
            file,
            folder="ems_avatars",
            transformation=[
                {'width': 200, 'height': 200, 'crop': 'fill', 'gravity': 'face'},
                {'fetch_format': 'auto', 'quality': 'auto'}
            ]
        )
        return res.get('secure_url')
    except Exception as e:
        app.logger.error(f"Cloudinary upload error: {e}")
        raise RuntimeError(f"Cloudinary integration failed: {str(e)}")

@app.route('/api/upload-avatar', methods=['POST'])
@login_required
def upload_avatar():
    if 'file' not in request.files:
        return jsonify({"success": False, "error": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "error": "No selected file"}), 400
    
    # 5MB size limit check
    file.seek(0, os.SEEK_END)
    file_length = file.tell()
    if file_length > 5 * 1024 * 1024:
        return jsonify({"success": False, "error": "File exceeds 5MB limit"}), 400
    file.seek(0) # reset pointer

    if not allowed_file(file.filename):
        return jsonify({"success": False, "error": "Invalid file type. Only JPG, PNG, WEBP allowed."}), 400

    # MIME type check
    allowed_mimes = {'image/png', 'image/jpeg', 'image/jpg', 'image/webp', 'image/pjpeg', 'image/x-png'}
    if not file.content_type or file.content_type.lower() not in allowed_mimes:
        return jsonify({"success": False, "error": "Invalid MIME type. Only JPG, PNG, WEBP allowed."}), 400

    target_type = request.form.get('target_type') # 'user', 'employee', 'part_time_worker'
    target_id = request.form.get('target_id')
    uid = get_current_user_id()

    if target_type not in ['user', 'employee', 'part_time_worker']:
        return jsonify({"success": False, "error": "Invalid target type"}), 400

    # Ownership checks
    if target_type == 'employee':
        emp_obj = safe_object_id(target_id)
        if not emp_obj:
            return jsonify({"success": False, "error": "Invalid Employee ID"}), 400
        emp = db.employees.find_one({"_id": emp_obj, "user_id": ObjectId(uid)})
        if not emp:
            return jsonify({"success": False, "error": "Unauthorized or Employee not found"}), 403
    elif target_type == 'part_time_worker':
        w_obj = safe_object_id(target_id)
        if not w_obj:
            return jsonify({"success": False, "error": "Invalid Worker ID"}), 400
        worker = db.part_time_workers.find_one({"_id": w_obj, "user_id": ObjectId(uid)})
        if not worker:
            return jsonify({"success": False, "error": "Unauthorized or Worker not found"}), 403

    try:
        # Upload to Cloudinary with transformations
        upload_result = cloudinary.uploader.upload(
            file,
            folder="ems_avatars",
            transformation=[
                {'width': 200, 'height': 200, 'crop': 'fill', 'gravity': 'face'},
                {'fetch_format': 'auto', 'quality': 'auto'}
            ]
        )
        image_url = upload_result.get('secure_url')

        # Update MongoDB
        if target_type == 'user':
            db.users.update_one({"_id": ObjectId(uid)}, {"$set": {"profile_image_url": image_url}})
            session['profile_image_url'] = image_url
            session.modified = True
        elif target_type == 'employee':
            db.employees.update_one({"_id": ObjectId(target_id), "user_id": ObjectId(uid)}, {"$set": {"profile_image_url": image_url}})
        elif target_type == 'part_time_worker':
            db.part_time_workers.update_one({"_id": ObjectId(target_id), "user_id": ObjectId(uid)}, {"$set": {"profile_image_url": image_url}})

        return jsonify({"success": True, "profile_image_url": image_url})
    except Exception as e:
        app.logger.error(f"Cloudinary upload error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/remove-avatar', methods=['POST'])
@login_required
def remove_avatar():
    data = request.json or {}
    target_type = data.get('target_type')
    target_id = data.get('target_id')
    uid = get_current_user_id()

    if target_type not in ['user', 'employee', 'part_time_worker']:
        return jsonify({"success": False, "error": "Invalid target type"}), 400

    # Ownership checks
    if target_type == 'employee':
        emp_obj = safe_object_id(target_id)
        if not emp_obj:
            return jsonify({"success": False, "error": "Invalid Employee ID"}), 400
        emp = db.employees.find_one({"_id": emp_obj, "user_id": ObjectId(uid)})
        if not emp:
            return jsonify({"success": False, "error": "Unauthorized or Employee not found"}), 403
    elif target_type == 'part_time_worker':
        w_obj = safe_object_id(target_id)
        if not w_obj:
            return jsonify({"success": False, "error": "Invalid Worker ID"}), 400
        worker = db.part_time_workers.find_one({"_id": w_obj, "user_id": ObjectId(uid)})
        if not worker:
            return jsonify({"success": False, "error": "Unauthorized or Worker not found"}), 403

    try:
        if target_type == 'user':
            db.users.update_one({"_id": ObjectId(uid)}, {"$unset": {"profile_image_url": ""}})
            session.pop('profile_image_url', None)
            session.modified = True
        elif target_type == 'employee':
            db.employees.update_one({"_id": ObjectId(target_id), "user_id": ObjectId(uid)}, {"$unset": {"profile_image_url": ""}})
        elif target_type == 'part_time_worker':
            db.part_time_workers.update_one({"_id": ObjectId(target_id), "user_id": ObjectId(uid)}, {"$unset": {"profile_image_url": ""}})

        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/export')
@limiter.limit("10 per minute")         # Excel export is expensive — cap it
@login_required
def export_data():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    uid = get_current_user_id()
    export_type = request.args.get('export_type', 'all_employees')
    export_type_alias_map = {
        'specific_part_time_worker': 'specific_part_time',
        'all_part_time_workers': 'all_part_time',
        'specific_employee_report': 'specific_employee',
        'all_employees_report': 'all_employees'
    }
    export_type = export_type_alias_map.get(export_type, export_type)
    export_format = request.args.get('format', 'excel')
    employee_id = request.args.get('employee_id')
    worker_id = request.args.get('worker_id')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # Validate date formats
    if start_date:
        try:
            datetime.strptime(start_date, '%Y-%m-%d')
        except ValueError:
            start_date = None
    if end_date:
        try:
            datetime.strptime(end_date, '%Y-%m-%d')
        except ValueError:
            end_date = None

    # Default date range to current month-to-date if not provided
    today_val = date.today()
    start_date_str = start_date or today_val.replace(day=1).isoformat()
    end_date_str = end_date or today_val.isoformat()

    try:
        d1 = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        d2 = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        if d1 > d2:
            d1, d2 = d2, d1
        delta = d2 - d1
        date_list = [(d1 + timedelta(days=i)).isoformat() for i in range(delta.days + 1)]
    except Exception:
        date_list = [today_val.isoformat()]

    # Styling Palettes
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Plus Jakarta Sans", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Plus Jakarta Sans", size=10)
    total_font = Font(name="Plus Jakarta Sans", size=11, bold=True)
    title_font = Font(name="Plus Jakarta Sans", size=16, bold=True, color="1E293B")

    pdf_sections = []
    
    def style_ws(ws, title=None, headers=None, rows=None, summary_row=None):
        if export_format == 'pdf':
            pdf_sections.append({
                'title': title,
                'headers': headers,
                'rows': rows,
                'summary_row': summary_row
            })
            return
            
        ws.views.sheetView[0].showGridLines = True
        current_row = 1
        if title:
            ws.cell(row=current_row, column=1, value=title).font = title_font
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers) if headers else 4)
            current_row += 2
            
        if headers:
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            ws.row_dimensions[current_row].height = 28
            current_row += 1
            
        if rows:
            for r_idx, r_data in enumerate(rows, current_row):
                for c_idx, val in enumerate(r_data, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = data_font
                    cell.border = thin_border
                    
                    if isinstance(val, (int, float)):
                        header_name = headers[c_idx - 1].lower() if headers and c_idx - 1 < len(headers) else ""
                        if any(term in header_name for term in ["salary", "amount", "advance", "deduction", "net", "price", "balance", "total"]):
                            cell.number_format = '₹#,##0.00'
                            cell.alignment = Alignment(horizontal="right")
                        else:
                            cell.number_format = '#,##0'
                            cell.alignment = Alignment(horizontal="right")
                    elif isinstance(val, (datetime, date)):
                        cell.number_format = 'YYYY-MM-DD'
                        cell.alignment = Alignment(horizontal="center")
                    elif isinstance(val, str) and len(val) == 10 and val.count('-') == 2:
                        cell.alignment = Alignment(horizontal="center")
                    else:
                        cell.alignment = Alignment(horizontal="left")
                ws.row_dimensions[r_idx].height = 20
                current_row += 1
                
        if summary_row:
            for c_idx, val in enumerate(summary_row, 1):
                if val is not None:
                    cell = ws.cell(row=current_row, column=c_idx, value=val)
                    cell.font = total_font
                    cell.border = thin_border
                    
                    header_name = headers[c_idx - 1].lower() if headers and c_idx - 1 < len(headers) else ""
                    if any(term in header_name for term in ["salary", "amount", "advance", "deduction", "net", "price", "balance", "total"]):
                        cell.number_format = '₹#,##0.00'
                    
                    if isinstance(val, str):
                        cell.alignment = Alignment(horizontal="left")
                    elif isinstance(val, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
            ws.row_dimensions[current_row].height = 24
            
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    val_str = str(cell.value)
                    if cell.number_format == '₹#,##0.00' and isinstance(cell.value, (int, float)):
                        val_str = f"Rs. {val_str}"
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb = openpyxl.Workbook()

    if export_type in ['all_employees', 'specific_employee']:
        # Fetch Employees
        emp_filter = {"user_id": ObjectId(uid)}
        if export_type == 'specific_employee' and employee_id:
            emp_id_obj = safe_object_id(employee_id)
            if emp_id_obj:
                emp_filter["_id"] = emp_id_obj
        
        emps = list(db.employees.find(emp_filter).sort("name", 1))
        emp_ids = [e['_id'] for e in emps]
        emp_map = {e['_id']: e for e in emps}
        months = sorted(list(set(dt[:7] for dt in date_list)))

        # 1. Fetch Attendance Logs (defaults to current month if dates are omitted)
        att_query = {"emp_id": {"$in": emp_ids}}
        q_start = start_date or start_date_str
        q_end = end_date or end_date_str
        att_query["date"] = {"$gte": q_start, "$lte": q_end}
        attendance_list = list(db.attendance.find(att_query))
        
        att_by_emp = {}
        for a in attendance_list:
            eid = str(a['emp_id'])
            if eid not in att_by_emp:
                att_by_emp[eid] = {}
            att_by_emp[eid][a['date']] = {
                'status': a.get('status', 'Absent'),
                'leave_reason': a.get('leave_reason', ''),
                'leave_note': a.get('leave_note', '')
            }

        # 2. Fetch Salary Records (all-time if no explicit dates, otherwise restricted to selected range)
        sal_query = {"emp_id": {"$in": emp_ids}}
        if start_date or end_date:
            sal_query["month"] = {"$in": months}
        salary_records_list = list(db.salary_records.find(sal_query))
        
        sal_by_emp = {}
        for s in salary_records_list:
            eid = str(s['emp_id'])
            m = s['month']
            if eid not in sal_by_emp:
                sal_by_emp[eid] = {}
            sal_by_emp[eid][m] = s

        # 3. Fetch Advances (all-time if no explicit dates, otherwise restricted to selected range)
        adv_query = {"emp_id": {"$in": emp_ids}}
        if start_date or end_date:
            adv_query["payment_date"] = {}
            if start_date:
                adv_query["payment_date"]["$gte"] = start_date
            if end_date:
                adv_query["payment_date"]["$lte"] = end_date + " 23:59:59"
        advances_list = list(db.salary_advance_payments.find(adv_query).sort("payment_date", -1))
        
        # 4. Fetch Advances for Summary calculation (must default to current month's advances if dates omitted)
        summary_adv_query = {"emp_id": {"$in": emp_ids}}
        s_start = start_date or start_date_str
        s_end = end_date or end_date_str
        summary_adv_query["payment_date"] = {"$gte": s_start, "$lte": s_end + " 23:59:59"}
        summary_advances_list = list(db.salary_advance_payments.find(summary_adv_query))
        
        adv_sum_by_emp = {}
        for adv in summary_advances_list:
            eid = str(adv['emp_id'])
            adv_sum_by_emp[eid] = adv_sum_by_emp.get(eid, 0.0) + adv['amount']

        # --- SHEET 1: SUMMARY ---
        ws_summary = wb.active
        ws_summary.title = "Summary"
        summary_headers = [
            "Employee ID", "Employee Name", "Department", "Phone Number", 
            "Joining Date", "Base Salary", "Present Days", "Absent Days", 
            "Leave Days", "Half Days", "Gross Salary", "Salary Advances", 
            "Deductions", "Net Salary", "Payment Status"
        ]
        
        summary_rows = []
        tot_base = tot_gross = tot_adv = tot_ded = tot_net = 0.0
        
        for e in emps:
            eid_str = str(e['_id'])
            p_days = a_days = l_days = h_days = 0
            
            for dt in date_list:
                status = att_by_emp.get(eid_str, {}).get(dt)
                if status == 'Absent':
                    a_days += 1
                elif status == 'Leave':
                    l_days += 1
                elif status == 'Half Day':
                    h_days += 1
                else:
                    p_days += 1
            
            base_sal = float(e.get('salary', 0.0))
            daily_rate = base_sal / 30.0
            # Gross based on present + half days
            gross = round(daily_rate * (p_days + (h_days * 0.5)), 2)
            adv_paid = round(adv_sum_by_emp.get(eid_str, 0.0), 2)
            deductions = adv_paid
            net = round(gross - deductions, 2)
            
            # Aggregate status
            statuses = [sal_by_emp.get(eid_str, {}).get(m, {}).get('payment_status', 'Pending') for m in months]
            if not statuses:
                p_status = 'Pending'
            elif 'Overpaid' in statuses:
                p_status = 'Overpaid'
            elif all(st == 'Settled' for st in statuses):
                p_status = 'Settled'
            else:
                p_status = 'Pending'
                
            tot_base += base_sal
            tot_gross += gross
            tot_adv += adv_paid
            tot_ded += deductions
            tot_net += net
            
            joining_date = e.get('joining_date') or (e.get('created_at')[:10] if e.get('created_at') else 'N/A')
            
            summary_rows.append([
                e.get('employee_id') or 'N/A',
                e.get('name', 'Unknown'),
                e.get('department', 'N/A'),
                e.get('phone', 'N/A'),
                joining_date,
                base_sal,
                p_days,
                a_days,
                l_days,
                h_days,
                gross,
                adv_paid,
                deductions,
                net,
                p_status
            ])
            
        summary_totals = [
            "TOTAL", "", "", "", "", tot_base, None, None, None, None, 
            tot_gross, tot_adv, tot_ded, tot_net, ""
        ]
        
        title_suffix = f" ({start_date_str} to {end_date_str})"
        style_ws(ws_summary, f"Employee Summary{title_suffix}", summary_headers, summary_rows, summary_totals)

        # --- SHEET 2: ATTENDANCE ---
        ws_attendance = wb.create_sheet(title="Attendance")
        att_headers = ["Employee ID", "Employee Name", "Date", "Status", "Reason", "Leave Note"]
        att_rows = []
        for e in emps:
            eid_str = str(e['_id'])
            for dt in date_list:
                att_data = att_by_emp.get(eid_str, {}).get(dt, {'status': 'Present', 'leave_reason': '', 'leave_note': ''})
                att_rows.append([
                    e.get('employee_id') or 'N/A',
                    e.get('name', 'Unknown'),
                    dt,
                    att_data.get('status', 'Present'),
                    att_data.get('leave_reason', ''),
                    att_data.get('leave_note', '')
                ])
        style_ws(ws_attendance, f"Attendance Logs{title_suffix}", att_headers, att_rows)

        # --- SHEET 3: SALARY ---
        ws_salary = wb.create_sheet(title="Salary")
        sal_headers = [
            "Employee ID", "Employee Name", "Month", "Base Salary", 
            "Present Days", "Gross Salary", "Total Advances", "Net Salary", "Payment Status"
        ]
        sal_rows = []
        # Determine the months to show in the Salary sheet
        if start_date or end_date:
            sal_months = months
        else:
            sal_months = sorted(list(set(s['month'] for s in salary_records_list)))
            if not sal_months:
                sal_months = [today_val.strftime('%Y-%m')]

        for e in emps:
            eid_str = str(e['_id'])
            for m in sal_months:
                rec = sal_by_emp.get(eid_str, {}).get(m, {})
                base_sal = float(e.get('salary', 0.0))
                p_days = rec.get('present_days', 30)
                gross = rec.get('total_salary', base_sal)
                adv = rec.get('advance_amount_paid', 0.0)
                net = round(gross - adv, 2)
                st = rec.get('payment_status', 'Pending')
                
                sal_rows.append([
                    e.get('employee_id') or 'N/A',
                    e.get('name', 'Unknown'),
                    m,
                    base_sal,
                    p_days,
                    gross,
                    adv,
                    net,
                    st
                ])
        sal_title_suffix = f" ({start_date} to {end_date})" if (start_date or end_date) else " (All Time)"
        style_ws(ws_salary, f"Monthly Salary Records{sal_title_suffix}", sal_headers, sal_rows)

        # --- SHEET 4: ADVANCES ---
        ws_advances = wb.create_sheet(title="Advances")
        adv_headers = ["Employee ID", "Employee Name", "Advance Amount", "Payment Date", "Notes", "Is Migrated"]
        adv_rows = []
        for adv in advances_list:
            emp = emp_map.get(adv['emp_id'], {})
            adv_rows.append([
                emp.get('employee_id') or 'N/A',
                emp.get('name', 'Unknown'),
                adv['amount'],
                adv.get('payment_date', 'N/A'),
                adv.get('notes', ''),
                adv.get('is_migrated', False)
            ])
        adv_title_suffix = f" ({start_date} to {end_date})" if (start_date or end_date) else " (All Time)"
        style_ws(ws_advances, f"Salary Advances Ledger{adv_title_suffix}", adv_headers, adv_rows)

    elif export_type in ['all_part_time', 'specific_part_time']:
        # Fetch Workers
        worker_filter = {"user_id": ObjectId(uid)}
        if export_type == 'specific_part_time' and worker_id:
            w_id_obj = safe_object_id(worker_id)
            if w_id_obj:
                worker_filter["_id"] = w_id_obj
                
        workers = list(db.part_time_workers.find(worker_filter).sort("name", 1))
        worker_ids = [w['_id'] for w in workers]
        worker_map = {w['_id']: w for w in workers}

        # Fetch Work Logs
        log_query = {"worker_id": {"$in": worker_ids}}
        if start_date or end_date:
            log_query["working_date"] = {}
            if start_date:
                log_query["working_date"]["$gte"] = start_date
            if end_date:
                log_query["working_date"]["$lte"] = end_date
        work_logs = list(db.part_time_work_logs.find(log_query).sort("working_date", -1))
        log_ids = [l['_id'] for l in work_logs]
        log_map = {l['_id']: l for l in work_logs}

        # Fetch Advances (use all logs of matching workers to ensure we capture all advances paid in this range)
        all_logs = list(db.part_time_work_logs.find({"worker_id": {"$in": worker_ids}}))
        all_log_ids = [l['_id'] for l in all_logs]
        all_log_map = {l['_id']: l for l in all_logs}
        log_worker_map = {l['_id']: l['worker_id'] for l in all_logs}
        
        # Fetch all-time advances for mapping remaining balances of logs accurately
        all_advances = list(db.advance_payments.find({
            "$or": [
                {"work_log_id": {"$in": all_log_ids}},
                {"worker_id": {"$in": worker_ids}}
            ]
        }))
        all_advances_by_log = {}
        for adv in all_advances:
            lid = adv.get('work_log_id')
            if lid:
                all_advances_by_log[lid] = all_advances_by_log.get(lid, 0.0) + adv['amount']
        
        # Fetch advances filtered by date range (for Advance History sheet & date range worker summary)
        adv_query = {
            "$or": [
                {"work_log_id": {"$in": all_log_ids}},
                {"worker_id": {"$in": worker_ids}}
            ]
        }
        if start_date or end_date:
            adv_query["payment_date"] = {}
            if start_date:
                adv_query["payment_date"]["$gte"] = start_date
            if end_date:
                adv_query["payment_date"]["$lte"] = end_date + " 23:59:59"
        advances = list(db.advance_payments.find(adv_query).sort("payment_date", -1))
        
        # Aggregate totals
        totals_by_worker = {}
        for w in workers:
            totals_by_worker[w['_id']] = {
                "total_work": 0.0,
                "total_adv": 0.0,
                "pending_count": 0
            }
            
        for log in work_logs:
            wid = log['worker_id']
            totals_by_worker[wid]["total_work"] += float(log.get('total_price', 0.0))
            if log.get('payment_status') == 'Pending':
                totals_by_worker[wid]["pending_count"] += 1
                
        for adv in advances:
            lid = adv.get('work_log_id')
            wid = adv.get('worker_id') or log_worker_map.get(lid)
            if wid:
                wid_obj = ObjectId(wid)
                if wid_obj in totals_by_worker:
                    totals_by_worker[wid_obj]["total_adv"] += adv['amount']

        # --- SHEET 1: WORKER SUMMARY ---
        ws_summary = wb.active
        ws_summary.title = "Worker Summary"
        worker_headers = [
            "Worker ID", "Worker Name", "Total Work Amount", 
            "Total Advances", "Remaining Balance", "Payment Status"
        ]
        worker_rows = []
        tot_work = tot_adv = tot_bal = 0.0
        
        for w in workers:
            stats = totals_by_worker[w['_id']]
            work_amt = stats["total_work"]
            adv_amt = stats["total_adv"]
            bal = round(work_amt - adv_amt, 2)
            
            p_status = 'Paid' if stats["pending_count"] == 0 and work_amt > 0 else 'Pending'
            if adv_amt > work_amt:
                p_status = 'Overpaid'
                
            tot_work += work_amt
            tot_adv += adv_amt
            tot_bal += bal
            
            worker_rows.append([
                w.get('worker_id') or 'N/A',
                w.get('name', 'Unknown'),
                work_amt,
                adv_amt,
                bal,
                p_status
            ])
            
        summary_totals = [
            "TOTAL", "", tot_work, tot_adv, tot_bal, ""
        ]
        pt_title_suffix = f" ({start_date} to {end_date})" if (start_date or end_date) else " (All Time)"
        style_ws(ws_summary, f"Part-Time Worker Summary{pt_title_suffix}", worker_headers, worker_rows, summary_totals)

        # --- SHEET 2: WORK LOGS ---
        ws_logs = wb.create_sheet(title="Work Logs")
        log_headers = [
            "Worker ID", "Worker Name", "Working Date", 
            "Client Name", "Delivery Location", "Slab Quantity", 
            "Price Per Slab", "Gross Amount", "Total Advances", "Remaining Balance", "Payment Status"
        ]
        log_rows = []
        pdf_log_rows = []
        for log in work_logs:
            w = worker_map.get(log['worker_id'], {})
            lid = log['_id']
            l_adv = all_advances_by_log.get(lid, 0.0)
            gross = float(log.get('total_price', 0.0))
            bal = round(gross - l_adv, 2)
            
            log_rows.append([
                w.get('worker_id') or 'N/A',
                w.get('name', 'Unknown'),
                log.get('working_date', 'N/A'),
                log.get('client_name', 'N/A'),
                log.get('delivery_location', 'N/A'),
                log.get('slab_quantity', 0),
                log.get('slab_price', 0.0),
                gross,
                l_adv,
                bal,
                log.get('payment_status', 'Pending')
            ])
            if export_type == 'specific_part_time':
                pdf_log_rows.append([
                    log.get('working_date', 'N/A'),
                    log.get('client_name', 'N/A'),
                    log.get('delivery_location', 'N/A'),
                    log.get('slab_quantity', 0),
                    log.get('slab_price', 0.0),
                    gross,
                    bal
                ])
            else:
                pdf_log_rows.append([
                    w.get('worker_id') or 'N/A',
                    w.get('name', 'Unknown'),
                    log.get('working_date', 'N/A'),
                    log.get('client_name', 'N/A'),
                    log.get('delivery_location', 'N/A'),
                    log.get('slab_quantity', 0),
                    log.get('slab_price', 0.0),
                    gross,
                    bal,
                    log.get('payment_status', 'Pending')
                ])

        if export_format == 'pdf':
            if export_type == 'specific_part_time':
                pdf_headers = ["Working Date", "Client Name", "Delivery Location", "Slab Quantity", "Price Per Slab", "Gross Amount", "Remaining Balance"]
            else:
                pdf_headers = ["Worker ID", "Worker Name", "Working Date", "Client Name", "Delivery Location", "Slab Quantity", "Price Per Slab", "Gross Amount", "Remaining Balance", "Payment Status"]
            style_ws(ws_logs, f"Part-Time Work Logs{pt_title_suffix}", pdf_headers, pdf_log_rows)
        else:
            style_ws(ws_logs, f"Part-Time Work Logs{pt_title_suffix}", log_headers, log_rows)

        # --- SHEET 3: ADVANCE HISTORY ---
        ws_advances = wb.create_sheet(title="Advance History")
        adv_headers = [
            "Worker ID", "Worker Name", "Reference", 
            "Advance Amount", "Advance Date", "Advance Notes"
        ]
        adv_rows = []
        for adv in advances:
            lid = adv.get('work_log_id')
            log = all_log_map.get(lid, {}) if lid else {}
            wid = adv.get('worker_id') or log.get('worker_id')
            w = worker_map.get(ObjectId(wid), {}) if wid else {}
            
            if lid and log:
                ref_str = f"Job (Client: {log.get('client_name', 'N/A')}, Date: {log.get('working_date', 'N/A')})"
            else:
                ref_str = "General Salary Advance"
                
            adv_rows.append([
                w.get('worker_id') or 'N/A',
                w.get('name', 'Unknown'),
                ref_str,
                adv['amount'],
                adv.get('payment_date', 'N/A'),
                adv.get('notes', '')
            ])
        style_ws(ws_advances, f"Part-Time Advances Ledger{pt_title_suffix}", adv_headers, adv_rows)

    elif export_type in ['all_attendance', 'specific_attendance']:
        # Fetch Employees
        emp_filter = {"user_id": ObjectId(uid)}
        if export_type == 'specific_attendance' and employee_id:
            emp_id_obj = safe_object_id(employee_id)
            if emp_id_obj:
                emp_filter["_id"] = emp_id_obj

        emps = list(db.employees.find(emp_filter).sort("name", 1))
        emp_ids = [e['_id'] for e in emps]

        # Fetch attendance within date range
        att_query = {"emp_id": {"$in": emp_ids}}
        att_query["date"] = {"$gte": start_date_str, "$lte": end_date_str}
        attendance_list = list(db.attendance.find(att_query).sort("date", 1))

        ws_att = wb.active
        ws_att.title = "Attendance"
        att_headers = ["Employee ID", "Employee Name", "Date", "Status", "Leave Reason", "Leave Note"]
        att_rows = []
        emp_map_att = {e['_id']: e for e in emps}

        for a in attendance_list:
            emp = emp_map_att.get(a['emp_id'], {})
            att_rows.append([
                emp.get('employee_id') or 'N/A',
                emp.get('name', 'Unknown'),
                a.get('date', 'N/A'),
                a.get('status', 'Present'),
                a.get('leave_reason', ''),
                a.get('leave_note', '')
            ])

        title_suffix = f" ({start_date_str} to {end_date_str})"
        style_ws(ws_att, f"Attendance Records{title_suffix}", att_headers, att_rows)

    elif export_type in ['all_salary', 'specific_salary']:
        # Fetch Employees
        emp_filter = {"user_id": ObjectId(uid)}
        if export_type == 'specific_salary' and employee_id:
            emp_id_obj = safe_object_id(employee_id)
            if emp_id_obj:
                emp_filter["_id"] = emp_id_obj

        emps = list(db.employees.find(emp_filter).sort("name", 1))
        emp_ids = [e['_id'] for e in emps]
        emp_map_sal = {e['_id']: e for e in emps}
        months = sorted(list(set(dt[:7] for dt in date_list)))

        # Fetch salary records
        sal_query = {"emp_id": {"$in": emp_ids}}
        if start_date or end_date:
            sal_query["month"] = {"$in": months}
        salary_records_list = list(db.salary_records.find(sal_query).sort("month", 1))

        # Fetch advances
        adv_query = {"emp_id": {"$in": emp_ids}}
        if start_date or end_date:
            adv_query["payment_date"] = {}
            if start_date:
                adv_query["payment_date"]["$gte"] = start_date
            if end_date:
                adv_query["payment_date"]["$lte"] = end_date + " 23:59:59"
        advances_list = list(db.salary_advance_payments.find(adv_query).sort("payment_date", -1))

        # Salary summary sheet
        ws_sal = wb.active
        ws_sal.title = "Salary Summary"
        sal_headers = [
            "Employee ID", "Employee Name", "Month", "Base Salary",
            "Present Days", "Gross Salary", "Total Advances", "Net Salary", "Payment Status"
        ]
        sal_rows = []
        tot_gross = tot_adv = tot_net = 0.0

        for rec in salary_records_list:
            emp = emp_map_sal.get(rec.get('emp_id'), {})
            base_sal = float(emp.get('salary', 0.0))
            p_days = rec.get('present_days', 30)
            gross = float(rec.get('total_salary', base_sal))
            adv = float(rec.get('advance_amount_paid', 0.0))
            net = round(gross - adv, 2)
            st = rec.get('payment_status', 'Pending')
            tot_gross += gross
            tot_adv += adv
            tot_net += net
            sal_rows.append([
                emp.get('employee_id') or 'N/A',
                emp.get('name', 'Unknown'),
                rec.get('month', 'N/A'),
                base_sal,
                p_days,
                gross,
                adv,
                net,
                st
            ])

        sal_title_suffix = f" ({start_date} to {end_date})" if (start_date or end_date) else " (All Time)"
        sal_totals = ["TOTAL", "", "", None, None, tot_gross, tot_adv, tot_net, ""]
        style_ws(ws_sal, f"Salary Records{sal_title_suffix}", sal_headers, sal_rows, sal_totals)

        # Advances sheet
        ws_adv = wb.create_sheet(title="Advances")
        adv_headers = ["Employee ID", "Employee Name", "Advance Amount", "Payment Date", "Notes"]
        adv_rows = []
        for adv in advances_list:
            emp = emp_map_sal.get(adv.get('emp_id'), {})
            adv_rows.append([
                emp.get('employee_id') or 'N/A',
                emp.get('name', 'Unknown'),
                float(adv.get('amount', 0.0)),
                adv.get('payment_date', 'N/A'),
                adv.get('notes', '')
            ])
        style_ws(ws_adv, f"Salary Advances{sal_title_suffix}", adv_headers, adv_rows)

    import tempfile, os
    from flask import Response
    
    fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    
    if export_format == 'pdf':
        try:
            os.remove(temp_path)
        except:
            pass
        from pdf_generator import generate_pdf_response
        return generate_pdf_response(pdf_sections, export_type)

    wb.save(temp_path)
    filename = f"export_{export_type}_{date.today().isoformat()}.xlsx"

    def generate_and_delete():
        try:
            with open(temp_path, 'rb') as f:
                while chunk := f.read(8192):
                    yield chunk
        finally:
            try:
                os.remove(temp_path)
            except Exception as e:
                app.logger.error(f"Failed to delete temp file {temp_path}: {e}")

    return Response(
        generate_and_delete(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
    )


# ─────────────────────────────────────────
# API - CHART DATA
# ─────────────────────────────────────────

@api_v1.route('/chart/attendance')
@limiter.limit("60 per minute")         # chart API called on month change
@login_required
def chart_attendance():
    uid = get_current_user_id()
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))

    emps = list(db.employees.find({"user_id": ObjectId(uid)}))
    labels, present_data, absent_data = [], [], []

    for e in emps:
        emp_id = e['_id']
        p = db.attendance.count_documents({
            "emp_id": emp_id,
            "status": "Present",
            "date": {"$regex": f"^{month}"}
        })
        a = db.attendance.count_documents({
            "emp_id": emp_id,
            "status": "Absent",
            "date": {"$regex": f"^{month}"}
        })
        labels.append(e.get('name', 'Unknown'))
        present_data.append(p)
        absent_data.append(a)

    return jsonify({'labels': labels, 'present': present_data, 'absent': absent_data})


# Register api_v1 blueprint after all its routes are declared
app.register_blueprint(api_v1)


# ─────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV', 'production') == 'development'
    print(f"Starting Employee Management System...")
    print(f"Open: http://127.0.0.1:{port}")
    app.run(host='0.0.0.0', port=port, debug=debug)
